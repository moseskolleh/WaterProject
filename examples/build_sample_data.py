"""Build the bundled sample datasets in the standard template formats.

The VES and pumping test values are transcribed from the two real
example documents (the Rokel geophysical survey report and the WiNGiN
field sheets). Lithology descriptions and the water quality results
are illustrative reconstructions so the full pipeline can be
exercised; see examples/README.md for what is verbatim and what is
filler.

Run from the repository root:

    python examples/build_sample_data.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from groundwater.ingestion.templates import (
    write_drilling_template,
    write_pumping_template,
    write_quality_template,
    write_ves_template,
)

HERE = Path(__file__).parent
DATA = HERE / "data"


# ---------------------------------------------------------------------------
# Rokel VES survey (transcribed from the survey report tables)
# ---------------------------------------------------------------------------

ROKEL_AB2 = [1, 2, 3, 3, 4, 5, 7, 10, 10, 15, 20, 30, 40, 40, 50, 70, 70, 80]
ROKEL_MN = [0.4, 0.4, 0.4, 0.8, 0.8, 0.8, 0.8, 0.8, 1.5, 1.5, 1.5, 1.5, 1.5, 7.6, 7.6, 7.6, 14, 14]
# apparent resistivities kept as strings exactly as printed, including
# leading zeros, to exercise the parser
ROKEL_RHO_A = ["1165", "1193", "1303", "1317", "1502", "1500", "1432", "1392",
               "961.0", "715.5", "732.0", "162.0", "156.1", "078.7", "052.1",
               "055.8", "053.2", "047.9"]
ROKEL_RHO_B = ["995.7", "984.2", "965.4", "1081", "1298", "1366", "1209", "1307",
               "891.8", "739.6", "329.1", "140.1", "147.9", "096.3", "034.1",
               "090.3", "055.2", "051.4"]

ROKEL_HEADERS = {
    "A (1)": {
        "client": "Living Water International",
        "community": "Rokel",
        "district": "Western Area",
        "sounding": "A (1)",
        "date": "8th December, 2015",
        "east": "0708958",
        "north": "0926355",
        "elevation": 71,
        "supervisor": "Field Supervisor",
        "rho": ROKEL_RHO_A,
    },
    "B (2)": {
        "client": "Living Water International",
        "community": "Rokel",
        # the original sheet carries this copy-over error on purpose: the
        # coordinates fall in the Western Area, the checker must flag it
        "district": "Port Loko",
        "sounding": "B (2)",
        "date": "8th December, 2015",
        "east": "0727012",
        "north": "0916125",
        "elevation": 68,
        "supervisor": "Field Supervisor",
        "rho": ROKEL_RHO_B,
    },
}

# IPI2Win models as printed in the report (N, rho, h, z, ERR)
ROKEL_IPI2WIN = {
    "A (1)": {"err": 21.5, "rows": [(1, 832.14, 1.0, "0/0"), (2, 2102.80, 7.37, 1.0), (3, 36.71, None, 8.37)]},
    "B (2)": {"err": 26.5, "rows": [(1, 1398.18, 0.71, "0/0"), (2, 703.00, 0.87, 0.71), (3, 1912.40, 8.42, 1.58), (4, 34.71, None, 10.00)]},
}


def build_rokel_ves() -> Path:
    folder = DATA / "rokel"
    folder.mkdir(parents=True, exist_ok=True)
    path = folder / "rokel_ves.xlsx"
    write_ves_template(path, n_soundings=2, n_rows=len(ROKEL_AB2))
    wb = load_workbook(path)
    for ws, key in zip(wb.worksheets, ("A (1)", "B (2)")):
        h = ROKEL_HEADERS[key]
        ws.title = key
        ws["B2"] = h["client"]; ws["D2"] = h["community"]
        ws["B3"] = "Geophysical Survey"; ws["D3"] = h["sounding"]
        ws["B4"] = h["district"]; ws["D4"] = h["east"]
        ws["B5"] = h["date"]; ws["D5"] = h["north"]
        ws["B6"] = h["supervisor"]; ws["D6"] = h["elevation"]
        ws["B7"] = ""; ws["D7"] = "28N"
        ws["B8"] = "Schlumberger"; ws["D8"] = "Syscal Junior"
        for i, (a, m, r) in enumerate(zip(ROKEL_AB2, ROKEL_MN, h["rho"])):
            row = 11 + i
            ws.cell(row=row, column=1, value=i + 1)
            ws.cell(row=row, column=2, value=a)
            ws.cell(row=row, column=3, value=m)
            ws.cell(row=row, column=4, value=r)  # string with leading zeros
    wb.save(path)
    return path


def build_rokel_ipi2win() -> Path:
    from openpyxl import Workbook

    folder = DATA / "rokel"
    folder.mkdir(parents=True, exist_ok=True)
    path = folder / "rokel_ipi2win_models.xlsx"
    wb = Workbook()
    first = True
    for sid, model in ROKEL_IPI2WIN.items():
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = sid
        ws["A1"] = "Sounding Number"; ws["B1"] = sid
        ws["A2"] = f"ERR = {model['err']}"
        ws["A4"] = "N"; ws["B4"] = "rho"; ws["C4"] = "h"; ws["D4"] = "z"
        for i, (n, rho, h, z) in enumerate(model["rows"]):
            r = 5 + i
            ws.cell(row=r, column=1, value=n)
            ws.cell(row=r, column=2, value=rho)
            ws.cell(row=r, column=3, value=h)
            ws.cell(row=r, column=4, value=z)
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Kuntolo step test (transcribed from the WiNGiN Word sheet)
# ---------------------------------------------------------------------------

KUNTOLO_STEP1 = [
    (1, 10.80), (2, 11.89), (3, 12.30), (4, 12.64), (5, 12.97), (6, 13.23),
    (7, 13.44), (8, 13.60), (9, 13.70), (10, 13.80), (12, 13.91), (14, 14.06),
    (16, 14.14), (18, 14.20), (20, 14.27), (22, 14.34), (24, 14.37), (26, 14.40),
    (28, 14.44), (30, 14.48), (32, 14.51), (34, 14.55), (36, 14.57), (38, 14.61),
    (40, 14.66), (42, 14.68), (44, 14.71), (46, 14.74), (48, 14.78), (50, 14.82),
    (52, 14.84), (55, 14.86), (57, 14.89), (60, 14.96),
]
KUNTOLO_STEP2 = [
    (61, 14.98), (62, 14.99), (63, 15.31), (64, 15.54), (65, 15.65), (66, 16.10),
    (67, 16.23), (68, 16.34), (69, 16.42), (70, 16.48), (72, 16.60), (74, 16.68),
    (76, 16.73), (78, 16.80), (80, 16.88), (82, 16.96), (84, 17.03), (86, 17.10),
    (88, 17.23), (90, 17.40), (92, 17.76), (94, 18.04), (96, 18.59), (98, 20.25),
    (100, 21.30), (102, 21.82), (104, 23.10), (106, 28.63), (108, 34.76),
    (110, 40.41), (112, 41.10), (115, 46.65), (117, 48.30), (120, 52.72),
]
KUNTOLO_STEP3 = [
    (121, 55.12), (122, 56.88), (123, 57.89), (124, 58.84), (125, 60.02),
    (126, 61.03), (127, 64.25), (128, 66.35), (129, 68.56), (130, 70.26),
    (132, 70.85), (134, 71.40), (136, 71.99), (138, 72.35), (140, 72.88),
    (142, 73.45), (144, 73.94), (146, 74.54), (148, 75.05), (150, 75.94),
    (152, 76.35), (154, 77.11), (156, 77.82), (158, 78.45),
]
KUNTOLO_RECOVERY = [
    (1, 67.56), (2, 63.26), (3, 60.00), (4, 59.98), (5, 59.50), (6, 59.32),
    (7, 58.94), (8, 58.57), (9, 58.14), (10, 57.83), (12, 57.06), (14, 56.28),
    (16, 55.40), (18, 54.68), (20, 53.94), (22, 53.15), (24, 52.44), (26, 51.91),
    (28, 50.88), (30, 50.06), (32, 49.31), (34, 48.62), (36, 47.95), (38, 47.22),
    (40, 46.28), (42, 45.62), (46, 45.12), (48, 44.16), (50, 43.56), (52, 42.99),
    (54, 42.02), (56, 40.87), (58, 40.16), (60, 39.49), (65, 37.78), (70, 35.67),
    (75, 34.17), (80, 30.82), (85, 29.13), (90, 27.66), (95, 23.31), (100, 17.18),
    (105, 13.96), (110, 10.39),
]


def build_kuntolo() -> Path:
    folder = DATA / "kuntolo"
    folder.mkdir(parents=True, exist_ok=True)
    path = folder / "kuntolo_step_test.xlsx"
    write_pumping_template(path)
    wb = load_workbook(path)
    ws = wb.active
    ws["B2"] = "Kuntoloh"; ws["E2"] = "1st April, 2017"
    ws["B3"] = "ACF"; ws["E3"] = 60
    ws["B4"] = "WiNGiN"; ws["E4"] = ""
    ws["B5"] = "KTL-01"; ws["E5"] = 70
    ws["B6"] = 19.28; ws["E6"] = 60
    ws["B7"] = "step"; ws["E7"] = "Port Loko"
    # discharge cells left empty on purpose: the sheet does not record them
    groups = [(1, KUNTOLO_STEP1), (4, KUNTOLO_STEP2), (7, KUNTOLO_STEP3)]
    start_row = 12
    for col0, series in groups:
        # incremental drawdown column written as recorded on paper
        prev = None
        for i, (t, wl) in enumerate(series):
            ws.cell(row=start_row + i, column=col0, value=t)
            ws.cell(row=start_row + i, column=col0 + 1, value=wl)
            inc = 0.0 if prev is None else round(wl - prev, 2)
            ws.cell(row=start_row + i, column=col0 + 2, value=inc)
            prev = wl
    prev = None
    for i, (t, wl) in enumerate(KUNTOLO_RECOVERY):
        ws.cell(row=start_row + i, column=13, value=t)
        ws.cell(row=start_row + i, column=14, value=wl)
        rec = 0.0 if prev is None else round(prev - wl, 2)
        ws.cell(row=start_row + i, column=15, value=rec)
        prev = wl
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Dr. Timbo completion data (drilling log + constant discharge test)
# ---------------------------------------------------------------------------

DR_TIMBO_PUMPING = [
    (0, 9.44), (1, 13.10), (2, 15.50), (3, 17.41), (4, 18.37), (5, 19.34),
    (10, 27.34), (15, 30.76), (20, 33.34), (25, 37.77), (30, 42.26),
]
DR_TIMBO_RECOVERY = [
    (0, 42.26), (1, 42.20), (2, 41.78), (3, 41.39), (4, 41.14), (5, 39.67),
    (10, 38.38), (15, 37.26), (20, 35.26), (25, 35.18), (30, 34.58), (35, 33.85),
    (40, 33.18), (45, 32.55), (50, 31.95), (55, 31.05), (60, 30.38),
]

# intervals verbatim from the drilling record; lithology descriptions are an
# illustrative reconstruction guided by the borehole diagram annotations
# ("Light Yellow Clayey Laterites", "Light Colour Granite", strikes at 12
# and 30 m)
DR_TIMBO_LOG = [
    ("0-5", "13:30", "13:35", 1.0, "Reddish brown lateritic topsoil", 6.5, None),
    ("5-10", "13:36", "13:41", 1.0, "Light yellow clayey laterites", 6.5, None),
    ("10-15", "13:55", "14:10", 0.33, "Light yellow clayey laterites, wet from 12 m", 6.5, 12),
    ("15-20", "14:14", "14:26", 0.42, "Clayey saprolite with weathered granite fragments", 6.5, None),
    ("20-25", "14:30", "14:38", 0.63, "Saprolite, decreasing clay content", 6.5, None),
    ("25-30", "15:15", "15:23", 0.63, "Weathered light colour granite, fractured", 6.5, 30),
    ("30-35", "15:25", "15:34", 0.56, "Light colour granite, slightly weathered, fractured", 6.5, None),
    ("35-40", "15:38", "15:48", 0.5, "Light colour granite, slightly weathered", 6.5, None),
    ("40-45", "15:50", "16:01", 0.45, "Light colour granite", 6.5, None),
    ("45-50", "16:03", "16:13", 0.5, "Light colour granite, fracture zone 49-52 m", 6.5, None),
    ("50-55", "16:15", "16:17", None, "Light colour granite", 6.5, None),
    ("55-60", "16:18", "16:20", None, "Light colour granite, fracture zone 60-62 m", 6.5, None),
    ("60-65", "16:27", "16:31", None, "Light colour granite", 6.5, None),
    ("65-70", "16:56", "17:00", None, "Light colour granite, fresh", 6.5, None),
]


def build_dr_timbo_drilling() -> Path:
    folder = DATA / "dr_timbo"
    folder.mkdir(parents=True, exist_ok=True)
    path = folder / "dr_timbo_drilling_log.xlsx"
    write_drilling_template(path, n_rows=len(DR_TIMBO_LOG))
    wb = load_workbook(path)
    ws = wb.active
    ws["B2"] = "Dr. Timbo's Residence"; ws["E2"] = "Dr. Timbo"
    ws["B3"] = "WiNGiN Heavy Duty Machines Co. Ltd"; ws["E3"] = "BH-1"
    ws["B4"] = "11th May, 2018"; ws["E4"] = "13th May, 2018"
    ws["B5"] = "Air rotary (DTH hammer)"; ws["E5"] = 70
    ws["B6"] = "Western Area Rural"; ws["E6"] = "Successful"
    ws["B7"] = ""; ws["E7"] = ""
    ws["B8"] = "28N"; ws["E8"] = ""
    ws["B9"] = 20; ws["E9"] = ""
    for i, row in enumerate(DR_TIMBO_LOG):
        r = 12 + i
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)
    wb.save(path)
    return path


def build_dr_timbo_pumping() -> Path:
    folder = DATA / "dr_timbo"
    folder.mkdir(parents=True, exist_ok=True)
    path = folder / "dr_timbo_constant_test.xlsx"
    write_pumping_template(path)
    wb = load_workbook(path)
    ws = wb.active
    ws["B2"] = "Dr. Timbo's Residence"; ws["E2"] = "10/05/2018"
    ws["B3"] = "Dr. Timbo"; ws["E3"] = ""
    ws["B4"] = "WiNGiN"; ws["E4"] = "17:14"
    ws["B5"] = "BH-1"; ws["E5"] = 70
    ws["B6"] = 9.44; ws["E6"] = 67
    ws["B7"] = "constant"; ws["E7"] = "Western Area Rural"
    ws["C9"] = 2.93  # discharge from the bucket measurement, m3/h
    start_row = 12
    prev = None
    for i, (t, wl) in enumerate(DR_TIMBO_PUMPING):
        ws.cell(row=start_row + i, column=1, value=t)
        ws.cell(row=start_row + i, column=2, value=wl)
        inc = 0.0 if prev is None else round(wl - prev, 2)
        ws.cell(row=start_row + i, column=3, value=inc)
        prev = wl
    prev = None
    for i, (t, wl) in enumerate(DR_TIMBO_RECOVERY):
        ws.cell(row=start_row + i, column=13, value=t)
        ws.cell(row=start_row + i, column=14, value=wl)
        rec = 0.0 if prev is None else round(prev - wl, 2)
        ws.cell(row=start_row + i, column=15, value=rec)
        prev = wl
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Water quality (synthetic, typical basement groundwater with a few
# exceedances to exercise the assessment)
# ---------------------------------------------------------------------------

QUALITY_VALUES = {
    "pH": 5.9,
    "Electrical conductivity": 185.0,
    "TDS": 120.0,
    "Turbidity": 3.2,
    "Temperature": 27.4,
    "Total hardness": 58.0,
    "Alkalinity": 46.0,
    "Calcium": 14.2,
    "Magnesium": 5.5,
    "Sodium": 12.6,
    "Potassium": 2.1,
    "Bicarbonate": 56.0,
    "Chloride": 18.4,
    "Sulfate": 6.8,
    "Nitrate (as NO3)": 8.2,
    "Nitrite (as NO2)": "<0.01",
    "Ammonia (as N)": 0.12,
    "Fluoride": 0.22,
    "Iron": 0.85,
    "Manganese": 0.12,
    "Arsenic": "<0.001",
    "Lead": "<0.005",
    "Copper": 0.02,
    "Zinc": 0.11,
    "Chromium (total)": "<0.01",
    "Cadmium": "<0.0005",
    "E. coli": 0,
    "Total coliforms": 4,
}


def build_dr_timbo_quality() -> Path:
    folder = DATA / "dr_timbo"
    folder.mkdir(parents=True, exist_ok=True)
    path = folder / "dr_timbo_water_quality.xlsx"
    write_quality_template(path)
    wb = load_workbook(path)
    ws = wb.active
    ws["B2"] = "Dr. Timbo's Residence"; ws["E2"] = "Dr. Timbo"
    ws["B3"] = "WQ-BH1-01"; ws["E3"] = "BH-1"
    ws["B4"] = "20th May, 2018"; ws["E4"] = "Example Laboratory, Freetown"
    ws["B5"] = "Western Area Rural"; ws["E5"] = "Borehole completion"
    for r in range(8, 8 + 40):
        name = ws.cell(row=r, column=1).value
        if name in QUALITY_VALUES:
            ws.cell(row=r, column=3, value=QUALITY_VALUES[name])
    wb.save(path)
    return path


def main() -> None:
    paths = [
        build_rokel_ves(),
        build_rokel_ipi2win(),
        build_kuntolo(),
        build_dr_timbo_drilling(),
        build_dr_timbo_pumping(),
        build_dr_timbo_quality(),
    ]
    for p in paths:
        print("wrote", p.relative_to(HERE.parent))


if __name__ == "__main__":
    main()
