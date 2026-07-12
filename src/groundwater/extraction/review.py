"""Review workbook generation for extracted documents.

Every extraction produces a review workbook: the header fields and
every table, with uncertain values highlighted amber and a Review
sheet listing everything that needs checking. For recognised VES
sheets the standard template can also be filled directly so the normal
parsers pick the data up after review.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from ..ingestion.common import match_label
from .models import REVIEW_THRESHOLD, ExtractedDocument

_AMBER = PatternFill("solid", fgColor="FFE08A")
_HEADER_FONT = Font(bold=True)


def write_review_workbook(document: ExtractedDocument, path: str | Path) -> Path:
    """Write the extraction to an Excel review workbook.

    Sheet 1 (Header): field, value, confidence; low confidence rows
    highlighted. One sheet per table with uncertain cells highlighted.
    Final Review sheet lists every flagged item with its reason.
    """
    path = Path(path)
    wb = Workbook()

    ws = wb.active
    ws.title = "Header"
    ws.append(["Field", "Value", "Confidence"])
    for cell in ws[1]:
        cell.font = _HEADER_FONT
    for field in document.header:
        ws.append([field.name, field.value, round(field.confidence, 2)])
        if field.needs_review:
            for cell in ws[ws.max_row]:
                cell.fill = _AMBER
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 32

    flagged = {(c.table_index, c.row, c.column): c.reason for c in document.uncertain_cells}
    for index, table in enumerate(document.tables):
        sheet = wb.create_sheet(f"Table {index + 1}"[:31])
        sheet.append([table.title])
        sheet["A1"].font = _HEADER_FONT
        sheet.append(table.columns)
        for cell in sheet[2]:
            cell.font = _HEADER_FONT
        for r, row in enumerate(table.rows):
            sheet.append(row)
            for c in range(len(row)):
                if (index, r, c) in flagged:
                    sheet.cell(row=sheet.max_row, column=c + 1).fill = _AMBER
            if table.confidence_for_row(r) < REVIEW_THRESHOLD:
                for c in range(len(row)):
                    sheet.cell(row=sheet.max_row, column=c + 1).fill = _AMBER

    review = wb.create_sheet("Review")
    review.append(["Items needing manual review"])
    review["A1"].font = _HEADER_FONT
    items = document.review_items
    if items:
        for item in items:
            review.append([item])
    else:
        review.append(["None: all values extracted with high confidence."])
    review.append([])
    review.append([f"Source: {document.source}"])
    review.append([f"Extractor: {document.extractor}"])
    review.append([f"Document kind: {document.document_kind}"])
    if document.notes:
        review.append([f"Notes: {document.notes}"])
    review.column_dimensions["A"].width = 100

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def fill_ves_template(document: ExtractedDocument, path: str | Path) -> Path:
    """Fill the standard VES template from an extracted VES sheet.

    Header fields are mapped through the same label patterns the
    parsers use; the first table supplies the readings. Values that
    need review keep their amber highlight in the filled template.
    """
    from ..ingestion.templates import write_ves_template

    if document.document_kind != "ves":
        raise ValueError("fill_ves_template needs a document of kind 'ves'")
    if not document.tables:
        raise ValueError("No data table was extracted")

    table = document.tables[0]
    path = Path(path)
    write_ves_template(path, n_soundings=1, n_rows=max(len(table.rows), 18))
    wb = load_workbook(path)
    ws = wb.active

    # header mapping: canonical key -> template value cell
    cell_map = {
        "client": "B2", "community": "D2",
        "project": "B3", "sounding_id": "D3",
        "district": "B4", "easting": "D4",
        "date": "B5", "northing": "D5",
        "supervisor": "B6", "elevation_m": "D6",
        "chiefdom": "B7", "utm_zone": "D7",
        "array_type": "B8", "instrument": "D8",
    }
    for field in document.header:
        key = match_label(field.name)
        target = cell_map.get(key or "")
        if target:
            ws[target] = field.value
            if field.needs_review:
                ws[target].fill = _AMBER

    flagged = {
        (c.row, c.column): c.reason
        for c in document.uncertain_cells
        if c.table_index == 0
    }
    # locate the AB/2, MN and resistivity columns in the extracted table
    def find_column(*needles):
        for i, name in enumerate(table.columns):
            low = name.lower()
            if any(n in low for n in needles):
                return i
        return None

    col_ab2 = find_column("ab/2", "ab2")
    col_mn = find_column("mn")
    col_rho = find_column("resist", "ohm", "rho")
    for r, row in enumerate(table.rows):
        target_row = 11 + r
        ws.cell(row=target_row, column=1, value=r + 1)
        for source_col, target_col in ((col_ab2, 2), (col_mn, 3), (col_rho, 4)):
            if source_col is None or source_col >= len(row):
                continue
            cell = ws.cell(row=target_row, column=target_col, value=row[source_col])
            if (r, source_col) in flagged:
                cell.fill = _AMBER
    wb.save(path)
    return path
