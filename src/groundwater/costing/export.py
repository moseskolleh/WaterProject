"""Bill of quantities Excel export.

Writes the estimate as a working BoQ workbook the contractor or client
can edit: quantities and rates stay as numbers with an amount formula
per line, so adjusting a rate updates the totals in Excel.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .model import CostEstimate

ACCENT = "1F5C8B"
LIGHT = "DCE6F1"

_thin = Side(style="thin", color="9BB3C8")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def write_boq_workbook(estimate: CostEstimate, path: str | Path) -> Path:
    """Write the bill of quantities with live amount formulas."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Bill of quantities"

    ws["A1"] = "BILL OF QUANTITIES - BOREHOLE CONSTRUCTION"
    ws["A1"].font = Font(bold=True, size=13, color=ACCENT)
    ws.merge_cells("A1:H1")
    ws["A2"] = (
        "Amounts are formulas (quantity x rate); edit quantities, rates or "
        "the assumptions block and every total updates. Rates in USD; the "
        "SLE column follows the exchange rate below."
    )
    ws["A2"].font = Font(italic=True, size=9)
    ws.merge_cells("A2:H2")

    headers = [
        "Code", "Stage", "Item", "Unit", "Quantity", "Rate (USD)",
        "Amount (USD)", "Amount (SLE)",
    ]
    header_row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=ACCENT)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BORDER

    row = header_row + 1
    first_item_row = row
    item_rows: list[int] = []
    current_stage = None
    for item in sorted(
        estimate.items,
        key=lambda i: (
            [s for s, _ in estimate.by_stage()].index(i.stage)
            if i.stage in dict(estimate.by_stage())
            else 99,
            i.code,
        ),
    ):
        if item.stage != current_stage:
            current_stage = item.stage
            cell = ws.cell(row=row, column=1, value=current_stage.upper())
            cell.font = Font(bold=True, size=10, color=ACCENT)
            cell.fill = PatternFill("solid", fgColor=LIGHT)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            row += 1
        item_rows.append(row)
        values = [
            item.code,
            item.stage,
            item.item,
            item.unit,
            round(item.quantity, 2),
            round(item.unit_cost_usd, 2),
            f"=E{row}*F{row}",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = BORDER
            if col in (5, 6, 7):
                cell.number_format = "#,##0.00"
        row += 1
    last_item_row = row - 1

    # ---- assumptions: the percentages a client negotiates, as cells -------
    # They used to be baked into the summary formulas as literals, so the
    # promise in row 2 ("edit ... and the totals update") stopped at the
    # direct cost. The summary below references these cells.
    row += 1
    cell = ws.cell(row=row, column=3, value="ASSUMPTIONS (edit; the summary follows)")
    cell.font = Font(bold=True, size=10, color=ACCENT)
    cell.fill = PatternFill("solid", fgColor=LIGHT)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
    row += 1
    assumption_cells: dict[str, str] = {}
    for key, label, value in (
        ("overheads", "Overheads (%)", estimate.overheads_percent),
        ("margin", "Margin (%)", estimate.margin_percent),
        ("vat", "VAT/GST (%)", estimate.vat_percent),
        ("contingency", "Contingency (%)", estimate.contingency_percent),
        ("fx", "Exchange rate (SLE per USD)", estimate.exchange_rate_sle_per_usd),
        ("depth", "Total depth drilled (m)", estimate.inputs.total_depth_m),
    ):
        ws.cell(row=row, column=3, value=label).font = Font(size=10)
        entry = ws.cell(row=row, column=5, value=value)
        entry.border = BORDER
        entry.fill = PatternFill("solid", fgColor="FFF8E1")
        entry.number_format = "0.00"
        assumption_cells[key] = f"$E${row}"
        row += 1
    fx = assumption_cells["fx"]
    for item_row in item_rows:
        sle = ws.cell(row=item_row, column=8, value=f"=G{item_row}*{fx}")
        sle.border = BORDER
        sle.number_format = "#,##0"

    def _summary(label: str, formula: str, bold: bool = False) -> str:
        nonlocal row
        cell = ws.cell(row=row, column=3, value=label)
        cell.font = Font(bold=True, size=10)
        amount = ws.cell(row=row, column=7, value=formula)
        amount.number_format = "#,##0.00"
        amount.font = Font(bold=bold, size=10)
        amount.border = BORDER
        local = ws.cell(row=row, column=8, value=f"=G{row}*{fx}")
        local.number_format = "#,##0"
        local.font = Font(bold=bold, size=10)
        local.border = BORDER
        row += 1
        return f"G{row - 1}"

    row += 1
    direct_cell = _summary(
        "Direct works cost",
        f"=SUMPRODUCT(E{first_item_row}:E{last_item_row},F{first_item_row}:F{last_item_row})",
        bold=True,
    )
    overheads_cell = _summary(
        "Overheads", f"={direct_cell}*{assumption_cells['overheads']}/100"
    )
    total_cell = _summary("Total cost", f"={direct_cell}+{overheads_cell}", bold=True)
    _summary("Cost per metre drilled", f"={total_cell}/{assumption_cells['depth']}")
    margin_cell = _summary("Margin", f"={total_cell}*{assumption_cells['margin']}/100")
    price_cell = _summary("Contract price", f"={total_cell}+{margin_cell}", bold=True)
    vat_cell = _summary("VAT/GST", f"={price_cell}*{assumption_cells['vat']}/100")
    with_vat_cell = _summary("Price including VAT", f"={price_cell}+{vat_cell}", bold=True)
    contingency_cell = _summary(
        "Contingency", f"={with_vat_cell}*{assumption_cells['contingency']}/100"
    )
    _summary("Planning budget", f"={with_vat_cell}+{contingency_cell}", bold=True)

    widths = (8, 14, 52, 10, 10, 12, 14, 14)
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(path)
    return path
