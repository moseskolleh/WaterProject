"""Shared helpers for reading field sheet layouts.

Field sheets use a header block of ``Label: value`` pairs followed by
one or more data tables. Labels vary between sheets ("Field
Supervisor", "Test conducted by", "Operator"), values sometimes live
in the same cell after a colon and sometimes several columns to the
right, and numbers may be stored as text with leading zeros. The
helpers here deal with all of that in one place.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from ..models import SiteMetadata
from ..utils import clean_text, parse_number

# Canonical header keys and the label patterns that map to them.
LABEL_PATTERNS: dict[str, list[str]] = {
    "client": [r"^client\b"],
    "project": [r"^project(?!\s*ref)"],
    "project_ref": [r"^project\s*ref", r"^geophysics\s*ref"],
    "community": [r"^community\b", r"^village\b", r"^location\b"],
    "chiefdom": [r"^chiefdom\b"],
    "district": [r"^district\b"],
    "sounding_id": [r"^sounding\s*(number|no)", r"^ves\s*(point|number|no)"],
    "date": [r"^date\b"],
    "supervisor": [
        r"^field\s*supervisor",
        r"^test\s*conducted\s*by",
        r"^operator\b",
        r"^supervisor\b",
    ],
    "contractor": [r"^(drilling\s*)?contractor\b"],
    "easting": [r"^gps\s*coordinate\s*east", r"^east(ing)?\b", r"^utm\s*east"],
    "northing": [r"^gps\s*coordinate\s*north", r"^north(ing)?\b", r"^utm\s*north"],
    "utm_zone": [r"^utm\s*zone", r"^zone\b"],
    "elevation_m": [r"^elevation", r"^altitude"],
    "array_type": [r"^array\b", r"^electrode\s*(array|configuration)"],
    "instrument": [r"^instrument\b", r"^equipment\b"],
    "borehole_ref": [r"^borehole\s*/?\s*ref", r"^bh\s*ref", r"^borehole\s*(number|no)"],
    "borehole_depth_m": [r"^depth\s*of\s*borehole", r"^borehole\s*depth", r"^total\s*depth", r"^depth\b"],
    "static_water_level_m": [r"^static\s*water\s*level", r"^swl\b"],
    "pump_setting_m": [r"^pump\s*setting", r"^pump\s*installation\s*depth"],
    "step_length_min": [r"^length\s*of\s*each\s*step", r"^step\s*length"],
    "test_type": [r"^test\s*type"],
    "start_time": [r"^time\b", r"^start\s*time"],
    "duration": [r"^duration"],
    "drilling_method": [r"^(drilling\s*)?method\b"],
    "start_date": [r"^(drilling\s*)?start\s*date"],
    "completion_date": [r"^compl(etion)?\.?\s*date"],
    "status": [r"^(bh\s*)?status"],
    "sample_id": [r"^sample\s*(id|number|no|ref)"],
    "laboratory": [r"^lab(oratory)?\b"],
    "sample_date": [r"^sample\s*date", r"^date\s*sampled"],
    "grouting_depth_m": [r"^grout(ing)?\b"],
    "drill_rig": [r"^drill\s*rig"],
}

_COMPILED = {
    key: [re.compile(p, re.IGNORECASE) for p in pats]
    for key, pats in LABEL_PATTERNS.items()
}

NUMERIC_HEADER_KEYS = {
    "easting",
    "northing",
    "elevation_m",
    "borehole_depth_m",
    "static_water_level_m",
    "pump_setting_m",
    "step_length_min",
    "grouting_depth_m",
}


def load_grid(path: str | Path, sheet: str | None = None) -> tuple[list[list], str]:
    """Load a worksheet as a rectangular grid of raw cell values."""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    grid = [list(row) for row in ws.iter_rows(values_only=True)]
    title = ws.title
    wb.close()
    return grid, title


def sheet_names(path: str | Path) -> list[str]:
    wb = load_workbook(path, read_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return names


def match_label(text: str) -> str | None:
    """Return the canonical key for a label cell, if any."""
    matched = match_label_priority(text)
    return matched[0] if matched else None


def match_label_priority(text: str) -> tuple[str, int] | None:
    """Return ``(key, priority)`` for a label cell.

    Priority is the pattern index within the key's list: 0 is the most
    specific wording ("Depth of Borehole"), higher numbers are generic
    fallbacks ("Depth"). A more specific match later in the sheet may
    overwrite a generic one.
    """
    label = clean_text(text).rstrip(":").strip()
    if not label:
        return None
    best: tuple[str, int] | None = None
    for key, patterns in _COMPILED.items():
        for priority, pattern in enumerate(patterns):
            if pattern.search(label):
                if best is None or priority < best[1]:
                    if priority == 0:
                        return key, 0
                    best = (key, priority)
                break
    return best


def split_inline_value(cell_text: str) -> tuple[str, str | None]:
    """Split ``"Community: Kuntoloh"`` into label and inline value."""
    text = clean_text(cell_text)
    if ":" in text:
        label, _, value = text.partition(":")
        value = value.strip()
        return label, value if value else None
    return text, None


def extract_header_fields(grid: list[list], max_rows: int = 30) -> dict:
    """Scan the top of a sheet for ``Label -> value`` pairs.

    A value is taken from the same cell after a colon when present,
    otherwise from the next non-empty cell to the right. The first
    occurrence of a key wins, except that a more specific label wording
    later in the sheet ("Depth of Borehole") overrides a generic one
    seen earlier ("Depth").
    """
    fields: dict = {}
    priorities: dict[str, int] = {}
    for r, row in enumerate(grid[:max_rows]):
        for c, cell in enumerate(row):
            if cell is None or isinstance(cell, (int, float)):
                continue
            label_text, inline_value = split_inline_value(str(cell))
            matched = match_label_priority(label_text)
            if matched is None:
                continue
            key, priority = matched
            if key in fields and priorities.get(key, 99) <= priority:
                continue
            value = inline_value
            if value is None:
                for cc in range(c + 1, min(c + 8, len(row))):
                    nxt = row[cc]
                    if nxt is not None and clean_text(nxt) != "":
                        # Stop if the neighbouring cell is itself a label
                        # (with or without its own inline value).
                        if not isinstance(nxt, (int, float)):
                            nxt_label, _ = split_inline_value(str(nxt))
                            if match_label(nxt_label):
                                break
                        value = nxt
                        break
            if value is None:
                continue
            if key in NUMERIC_HEADER_KEYS:
                number = parse_number(value)
                if number is not None:
                    fields[key] = number
                    priorities[key] = priority
            else:
                fields[key] = clean_text(value)
                priorities[key] = priority
    return fields


def site_from_fields(fields: dict, source: str = "") -> SiteMetadata:
    """Build SiteMetadata from extracted header fields."""
    zone = fields.get("utm_zone")
    if isinstance(zone, str):
        zone = parse_number(zone)
    return SiteMetadata(
        client=fields.get("client", ""),
        project=fields.get("project", ""),
        community=fields.get("community", ""),
        chiefdom=fields.get("chiefdom", ""),
        district=fields.get("district", ""),
        project_ref=fields.get("project_ref", ""),
        easting=fields.get("easting"),
        northing=fields.get("northing"),
        utm_zone=int(zone) if zone else None,
        elevation_m=fields.get("elevation_m"),
        date=str(fields.get("date", "")),
        supervisor=fields.get("supervisor", ""),
        contractor=fields.get("contractor", ""),
        source=source,
    )


def find_row(grid: list[list], predicate, start: int = 0) -> int | None:
    """Index of the first row for which ``predicate(row)`` is true."""
    for r in range(start, len(grid)):
        if predicate(grid[r]):
            return r
    return None


def row_text(row: list) -> list[str]:
    return [clean_text(c).lower() for c in row]
