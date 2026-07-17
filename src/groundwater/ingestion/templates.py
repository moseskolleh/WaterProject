"""Generate the standard Excel field data templates.

The templates mirror the paper field sheets in use (Rokel VES sheet,
WiNGiN step test sheet) so the field team can transcribe data cell for
cell, while adding the fields that are often missing on paper, such as
discharge per step and the UTM zone.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ACCENT = "1F5C8B"
LIGHT = "DCE6F1"

_thin = Side(style="thin", color="9BB3C8")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _label(ws, cell: str, text: str) -> None:
    ws[cell] = text
    ws[cell].font = Font(bold=True, size=10)
    ws[cell].fill = PatternFill("solid", fgColor=LIGHT)
    ws[cell].border = BORDER


def _value(ws, cell: str, text: str = "") -> None:
    ws[cell] = text
    ws[cell].border = BORDER


def _title(ws, cell: str, text: str, span: str | None = None) -> None:
    ws[cell] = text
    ws[cell].font = Font(bold=True, size=12, color=ACCENT)
    if span:
        ws.merge_cells(span)
        ws[cell].alignment = Alignment(horizontal="center")


def _table_header(ws, row: int, headers: list[str], start_col: int = 1) -> None:
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=h)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=ACCENT)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BORDER


def write_ves_template(path: str | Path, n_soundings: int = 2, n_rows: int = 24) -> Path:
    """VES field data template, one worksheet per sounding."""
    wb = Workbook()
    for s in range(1, n_soundings + 1):
        ws = wb.active if s == 1 else wb.create_sheet()
        ws.title = f"VES {s}"
        _title(ws, "A1", "SCHLUMBERGER ARRAY VES FIELD DATA", "A1:D1")
        pairs = [
            ("A2", "Client", "B2"), ("C2", "Community", "D2"),
            ("A3", "Project", "B3"), ("C3", "Sounding Number", "D3"),
            ("A4", "District", "B4"), ("C4", "GPS Coordinate East", "D4"),
            ("A5", "Date", "B5"), ("C5", "GPS Coordinate North", "D5"),
            ("A6", "Field Supervisor", "B6"), ("C6", "Elevation (m)", "D6"),
            ("A7", "Chiefdom", "B7"), ("C7", "UTM Zone (28N or 29N)", "D7"),
            ("A8", "Array", "B8"), ("C8", "Instrument", "D8"),
        ]
        for lab_cell, lab, val_cell in pairs:
            _label(ws, lab_cell, lab)
            _value(ws, val_cell)
        ws["B8"] = "Schlumberger"
        ws["D8"] = "Syscal Junior"
        ws["D3"] = s
        _table_header(ws, 10, ["No.", "AB/2 (m)", "MN (m)", "Apparent Resistivity (ohm-m)"])
        for r in range(11, 11 + n_rows):
            ws.cell(row=r, column=1, value=r - 10).border = BORDER
            for c in range(2, 5):
                ws.cell(row=r, column=c).border = BORDER
        note = (
            "Notes: MN is the full potential electrode spacing. Repeat the same AB/2 "
            "with the old and new MN at every segment change."
        )
        ws.cell(row=12 + n_rows, column=1, value=note).font = Font(italic=True, size=9)
        for col, width in zip("ABCD", (22, 16, 16, 28)):
            ws.column_dimensions[col].width = width
    wb.save(path)
    return Path(path)


def write_pumping_template(path: str | Path) -> Path:
    """Pumping test template: header, four step groups and recovery."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pumping Test"
    _title(ws, "A1", "PUMPING TEST FIELD SHEET (STEP / CONSTANT DISCHARGE)", "A1:O1")
    pairs = [
        ("A2", "Community", "B2"), ("D2", "Date", "E2"),
        ("A3", "Client", "B3"), ("D3", "Length of each step (min)", "E3"),
        ("A4", "Test conducted by", "B4"), ("D4", "Start time", "E4"),
        ("A5", "Borehole Ref. No.", "B5"), ("D5", "Depth of Borehole (m)", "E5"),
        ("A6", "Static water level (m)", "B6"), ("D6", "Pump setting (m)", "E6"),
        ("A7", "Test type (step or constant)", "B7"), ("D7", "District", "E7"),
        ("G2", "GPS Coordinate East", "H2"), ("G3", "GPS Coordinate North", "H3"),
        ("G4", "UTM Zone (28N or 29N)", "H4"), ("G5", "Elevation (m)", "H5"),
    ]
    for lab_cell, lab, val_cell in pairs:
        _label(ws, lab_cell, lab)
        _value(ws, val_cell)

    # Discharge row: one value per step group; leave blank if not measured
    _label(ws, "A9", "Discharge per step (m3/h)")
    for i, col in enumerate(("B", "E", "H", "K")):
        _label(ws, f"{col}9", f"Step {i + 1} Q")
        _value(ws, f"{chr(ord(col) + 1)}9")

    groups = ["0-60 min", "61-120 min", "121-180 min", "181-240 min"]
    start_cols = [1, 4, 7, 10]
    header_row = 11
    for g, (label, c0) in enumerate(zip(groups, start_cols)):
        cell = ws.cell(row=header_row - 1, column=c0, value=f"Constant discharge {label}")
        cell.font = Font(bold=True, size=10, color=ACCENT)
        _table_header(
            ws, header_row, ["Time (min)", "Water Level (m)", "Drawdown (m)"], start_col=c0
        )
    cell = ws.cell(row=header_row - 1, column=13, value="Recovery")
    cell.font = Font(bold=True, size=10, color=ACCENT)
    _table_header(
        ws, header_row, ["Time (min)", "Water Level (m)", "Recovery (m)"], start_col=13
    )
    for r in range(header_row + 1, header_row + 41):
        for c in list(range(1, 12 + 1)) + [13, 14, 15]:
            ws.cell(row=r, column=c).border = BORDER
    note = (
        "Notes: record depth to water in metres below the measuring point. The Drawdown "
        "and Recovery columns are the change between successive readings; the analysis "
        "recomputes true drawdown from the static water level. Recovery time is minutes "
        "since the pump stopped. Record the discharge of every step; results stay "
        "provisional until discharge is supplied."
    )
    ws.cell(row=header_row + 42, column=1, value=note).font = Font(italic=True, size=9)
    for c in range(1, 16):
        ws.column_dimensions[get_column_letter(c)].width = 13
    wb.save(path)
    return Path(path)


def write_drilling_template(path: str | Path, n_rows: int = 20) -> Path:
    """Drilling log template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Drilling Log"
    _title(ws, "A1", "BOREHOLE DRILLING LOG", "A1:G1")
    pairs = [
        ("A2", "Community", "B2"), ("D2", "Client", "E2"),
        ("A3", "Contractor", "B3"), ("D3", "Borehole Ref. No.", "E3"),
        ("A4", "Drilling start date", "B4"), ("D4", "Completion date", "E4"),
        ("A5", "Drilling method", "B5"), ("D5", "Total depth (m)", "E5"),
        ("A6", "District", "B6"), ("D6", "BH status", "E6"),
        ("A7", "GPS Coordinate East", "B7"), ("D7", "GPS Coordinate North", "E7"),
        ("A8", "UTM Zone (28N or 29N)", "B8"), ("D8", "Elevation (m)", "E8"),
        ("A9", "Grouting depth (m)", "B9"), ("D9", "Drill rig", "E9"),
    ]
    for lab_cell, lab, val_cell in pairs:
        _label(ws, lab_cell, lab)
        _value(ws, val_cell)
    _table_header(
        ws,
        11,
        [
            "Depth interval (m)",
            "From time",
            "To time",
            "Penetration rate (m/min)",
            "Sample / lithology description",
            "Drilling diameter (in)",
            "Water strike depth (m)",
        ],
    )
    for r in range(12, 12 + n_rows):
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = BORDER
    ws.cell(row=13 + n_rows, column=1, value=(
        "Notes: write depth intervals as 0-5, 5-10 and so on. Enter a water strike "
        "depth on the row of the interval where water was struck."
    )).font = Font(italic=True, size=9)
    for col, width in zip("ABCDEFG", (18, 10, 10, 20, 42, 18, 20)):
        ws.column_dimensions[col].width = width
    wb.save(path)
    return Path(path)


def write_daily_log_template(path: str | Path, n_rows: int = 14) -> Path:
    """Driller's daily report template.

    The daily record the supervision guidance expects: per interval
    times, depths, formation, water strikes and airlift yield, plus
    the day totals, standing time and the double signature (rig
    operator and supervisor) that makes the log auditable against
    invoiced metres.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Log"
    _title(ws, "A1", "DRILLER'S DAILY REPORT", "A1:G1")
    pairs = [
        ("A2", "Community", "B2"), ("D2", "Borehole Ref. No.", "E2"),
        ("A3", "Date", "B3"), ("D3", "Drill rig", "E3"),
        ("A4", "Contractor", "B4"), ("D4", "Supervisor", "E4"),
        ("A5", "Weather / site conditions", "B5"), ("D5", "Record taker", "E5"),
    ]
    for lab_cell, lab, val_cell in pairs:
        _label(ws, lab_cell, lab)
        _value(ws, val_cell)
    _table_header(
        ws,
        7,
        [
            "Time from",
            "Time to",
            "Depth from (m)",
            "Depth to (m)",
            "Formation / activity",
            "Water strike (m)",
            "Airlift yield (L/s)",
        ],
    )
    for r in range(8, 8 + n_rows):
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = BORDER
    totals_row = 9 + n_rows
    for cell, label in (
        (f"A{totals_row}", "Metres drilled today"),
        (f"C{totals_row}", "Cumulative metres"),
        (f"E{totals_row}", "Casing installed today (m)"),
        (f"A{totals_row + 1}", "Standing / breakdown hours"),
        (f"C{totals_row + 1}", "Reason"),
    ):
        _label(ws, cell, label)
    for cell in (f"B{totals_row}", f"D{totals_row}", f"F{totals_row}",
                 f"B{totals_row + 1}", f"D{totals_row + 1}"):
        _value(ws, cell)
    sign_row = totals_row + 3
    for cell, label in (
        (f"A{sign_row}", "Rig operator signature"),
        (f"D{sign_row}", "Supervisor signature"),
    ):
        _label(ws, cell, label)
        _value(ws, cell.replace("A", "B").replace("D", "E"))
    ws.cell(row=sign_row + 2, column=1, value=(
        "Notes: one row per drilled interval or activity (moving, standing, "
        "casing). Both signatures are required every day; the office checks "
        "invoiced metres against these logs."
    )).font = Font(italic=True, size=9)
    for col, width in zip("ABCDEFG", (12, 12, 14, 14, 40, 16, 16)):
        ws.column_dimensions[col].width = width
    wb.save(path)
    return Path(path)


STANDARD_PARAMETERS = [
    ("pH", "pH units"), ("Electrical conductivity", "uS/cm"), ("TDS", "mg/L"),
    ("Turbidity", "NTU"), ("Temperature", "deg C"), ("Total hardness", "mg/L as CaCO3"),
    ("Alkalinity", "mg/L as CaCO3"), ("Calcium", "mg/L"), ("Magnesium", "mg/L"),
    ("Sodium", "mg/L"), ("Potassium", "mg/L"), ("Bicarbonate", "mg/L"),
    ("Chloride", "mg/L"), ("Sulfate", "mg/L"), ("Nitrate (as NO3)", "mg/L"),
    ("Nitrite (as NO2)", "mg/L"), ("Ammonia (as N)", "mg/L"), ("Fluoride", "mg/L"),
    ("Iron", "mg/L"), ("Manganese", "mg/L"), ("Arsenic", "mg/L"), ("Lead", "mg/L"),
    ("Copper", "mg/L"), ("Zinc", "mg/L"), ("Chromium (total)", "mg/L"),
    ("Cadmium", "mg/L"), ("E. coli", "CFU/100 mL"), ("Total coliforms", "CFU/100 mL"),
]


def write_quality_template(path: str | Path) -> Path:
    """Water quality laboratory results template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Water Quality"
    _title(ws, "A1", "WATER QUALITY LABORATORY RESULTS", "A1:E1")
    pairs = [
        ("A2", "Community", "B2"), ("D2", "Client", "E2"),
        ("A3", "Sample ID", "B3"), ("D3", "Borehole Ref. No.", "E3"),
        ("A4", "Sample date", "B4"), ("D4", "Laboratory", "E4"),
        ("A5", "District", "B5"), ("D5", "Project", "E5"),
    ]
    for lab_cell, lab, val_cell in pairs:
        _label(ws, lab_cell, lab)
        _value(ws, val_cell)
    _table_header(
        ws, 7, ["Parameter", "Unit", "Value", "Detection limit", "Method"]
    )
    for i, (param, unit) in enumerate(STANDARD_PARAMETERS):
        r = 8 + i
        ws.cell(row=r, column=1, value=param).border = BORDER
        ws.cell(row=r, column=2, value=unit).border = BORDER
        for c in (3, 4, 5):
            ws.cell(row=r, column=c).border = BORDER
    ws.cell(row=9 + len(STANDARD_PARAMETERS), column=1, value=(
        "Notes: for results below detection write <DL in Value (for example <0.01) "
        "or leave Value empty and fill Detection limit. Add extra parameters on new rows."
    )).font = Font(italic=True, size=9)
    for col, width in zip("ABCDE", (26, 16, 12, 16, 18)):
        ws.column_dimensions[col].width = width
    wb.save(path)
    return Path(path)


def write_all_templates(folder: str | Path) -> list[Path]:
    """Write every field template into ``folder`` and return the paths."""
    folder = Path(folder)
    folder.mkdir(parents=True, exist_ok=True)
    return [
        write_ves_template(folder / "template_ves.xlsx"),
        write_pumping_template(folder / "template_pumping_test.xlsx"),
        write_drilling_template(folder / "template_drilling_log.xlsx"),
        write_quality_template(folder / "template_water_quality.xlsx"),
        write_daily_log_template(folder / "template_daily_drilling_report.xlsx"),
    ]
