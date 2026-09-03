import numpy as np

import pytest

from groundwater.ingestion import (
    read_drilling_workbook,
    read_pumping_workbook,
    read_quality_workbook,
    read_ves_workbook,
)


def test_ves_parsing(sample_data):
    soundings = read_ves_workbook(sample_data / "rokel" / "rokel_ves.xlsx")
    assert len(soundings) == 2
    a = soundings[0]
    assert a.sounding_id == "A (1)"
    assert a.n_readings == 18
    # leading-zero strings parse as numbers
    assert a.rho_app[13] == 78.7
    assert a.site.easting == 708958
    assert a.site.northing == 926355
    # duplicate AB/2 at segment changes preserved
    assert np.sum(a.ab2 == 40) == 2
    assert any(f.code == "segment_overlap" for f in a.flags)
    # second sounding carries the copy-over district error from the source
    assert soundings[1].site.district == "Port Loko"


def test_pumping_parsing_kuntolo(sample_data):
    test = read_pumping_workbook(sample_data / "kuntolo" / "kuntolo_step_test.xlsx")
    assert test.test_type.startswith("step")
    assert len(test.steps) == 3
    assert test.static_water_level_m == 19.28
    assert test.step_length_min == 60
    # irregular time spacing preserved
    assert list(test.steps[0].time_min[:3]) == [1.0, 2.0, 3.0]
    assert 55.0 in test.steps[0].time_min  # 52 -> 55 jump
    # incremental drawdown column ignored; water levels kept
    assert test.steps[0].water_level_m[0] == 10.80
    # recovery block read from its own column group
    assert test.recovery_time_min is not None
    assert len(test.recovery_time_min) == 44
    # discharge missing -> flagged as pending
    assert not test.has_discharge
    assert any(f.code == "missing_discharge" for f in test.flags)
    # the negative-drawdown anomaly on this sheet is flagged
    assert any(f.code == "water_level_above_static" for f in test.flags)


def test_pumping_parsing_dr_timbo(sample_data):
    test = read_pumping_workbook(sample_data / "dr_timbo" / "dr_timbo_constant_test.xlsx")
    assert test.test_type.startswith("constant")
    assert len(test.steps) == 1
    assert test.steps[0].discharge_m3_per_h == 2.93
    assert test.static_water_level_m == 9.44
    assert test.recovery_time_min is not None
    # true drawdown recomputed from static level, not the increment column
    drawdown = test.drawdown(test.steps[0])
    assert abs(drawdown[-1] - (42.26 - 9.44)) < 1e-9


def test_drilling_parsing(sample_data):
    log = read_drilling_workbook(sample_data / "dr_timbo" / "dr_timbo_drilling_log.xlsx")
    assert log.total_depth_m == 70
    assert len(log.intervals) == 14
    assert log.intervals[0].top_m == 0 and log.intervals[0].bottom_m == 5
    assert log.water_strikes_m == [12.0, 30.0]
    assert log.grouting_depth_m == 20
    assert not any(f.code == "interval_overlap" for f in log.flags)


def test_quality_parsing(sample_data):
    sample = read_quality_workbook(sample_data / "dr_timbo" / "dr_timbo_water_quality.xlsx")
    assert len(sample.results) >= 25
    iron = sample.get("Iron")
    assert iron is not None and iron.value == 0.85
    nitrite = sample.get("Nitrite (as NO2)")
    assert nitrite.below_detection and nitrite.detection_limit == 0.01


def test_an_excel_date_cell_in_the_header_reads_as_a_date_not_a_timestamp():
    """openpyxl hands a date-typed cell back as a datetime; str() of that put
    'Survey date: 2015-12-08 00:00:00' on the report cover. ISO date is what
    the browser app prints for the same cell."""
    import datetime

    from groundwater.ingestion.common import extract_header_fields, site_from_fields

    grid = [
        ["Client", "Living Water", None, "Date", datetime.datetime(2015, 12, 8)],
        ["Community", "Rokel", None, "District", "Western Area"],
    ]
    fields = extract_header_fields(grid)
    assert fields["date"] == "2015-12-08"
    assert site_from_fields(fields).date == "2015-12-08"
    grid[0][4] = datetime.date(2015, 12, 8)
    assert extract_header_fields(grid)["date"] == "2015-12-08"
    grid[0][4] = "8th December, 2015"
    assert extract_header_fields(grid)["date"] == "8th December, 2015"


def test_a_columnar_header_block_is_read_from_the_row_beneath():
    """Labels across one row and values in the next used to parse to nothing:
    the cover printed blanks and the consistency check reported the
    coordinates missing although they were on the sheet."""
    from groundwater.ingestion.common import extract_header_fields

    grid = [
        ["Client", "Community", "District", "Sounding Number",
         "GPS Coordinate East", "GPS Coordinate North"],
        ["Living Water", "Rokel", "Western Area", "VES 2", 708958, 926355],
        [],
        ["No.", "AB/2 (m)", "MN (m)", "Rho (ohm.m)"],
    ]
    fields = extract_header_fields(grid)
    assert fields["community"] == "Rokel"
    assert fields["district"] == "Western Area"
    assert fields["easting"] == 708958 and fields["northing"] == 926355
    assert fields["sounding_id"] == "VES 2"
    # the stacked layout is unchanged: a label over a label is not a value
    stacked = [["Client", "Living Water"], ["Community", "Rokel"], ["District", None]]
    assert extract_header_fields(stacked)["community"] == "Rokel"
    assert "district" not in extract_header_fields(stacked)


def _ves_workbook(path, header, rows, sheet="VES 1"):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(["Client", "Living Water"])
    ws.append(["Community", "Rokel"])
    ws.append(["Sounding Number", sheet])
    ws.append([])
    ws.append(header)
    for row in rows:
        ws.append(row)
    wb.save(path)
    return path


def test_greek_resistivity_headers_are_read(tmp_path):
    """'ρa (Ω·m)' is how a geophysicist labels the column; the reader used
    to drop the sheet without a word."""
    from groundwater.ingestion import read_ves_workbook

    path = _ves_workbook(
        tmp_path / "greek.xlsx", ["No.", "AB/2 (m)", "MN (m)", "ρa (Ω·m)"],
        [[1, 1, 0.4, 1165], [2, 2, 0.4, 1193], [3, 3, 0.4, 1303], [4, 5, 0.4, 1500]],
    )
    soundings = read_ves_workbook(path)
    assert len(soundings) == 1
    assert list(soundings[0].rho_app) == [1165, 1193, 1303, 1500]


def test_a_resistance_column_is_not_a_resistivity(tmp_path):
    """A sheet that records V/I in ohms matched the 'ohm' test and every
    reading came through as a resistivity of 0.9; now rho = K x R."""
    from groundwater.ingestion import read_ves_workbook
    from groundwater.ves.arrays import geometric_factor

    path = _ves_workbook(
        tmp_path / "resistance.xlsx", ["No.", "AB/2 (m)", "MN (m)", "R (ohm)"],
        [[1, 1, 0.4, 150.0], [2, 2, 0.4, 40.0], [3, 3, 0.4, 18.0], [4, 5, 0.4, 6.5]],
    )
    soundings = read_ves_workbook(path)
    assert len(soundings) == 1
    k = geometric_factor("schlumberger", ab2=2.0, mn=0.4)
    assert soundings[0].rho_app[1] == pytest.approx(float(k) * 40.0)
    assert any(f.code == "rho_computed_from_resistance" for f in soundings[0].flags)


def test_a_skipped_sheet_is_named_with_its_reason(tmp_path):
    from openpyxl import Workbook

    from groundwater.ingestion import read_ves_workbook

    path = _ves_workbook(
        tmp_path / "mixed.xlsx", ["No.", "AB/2 (m)", "MN (m)", "Rho (ohm.m)"],
        [[1, 1, 0.4, 1165], [2, 2, 0.4, 1193], [3, 3, 0.4, 1303], [4, 5, 0.4, 1500]],
    )
    from openpyxl import load_workbook

    wb = load_workbook(path)
    bad = wb.create_sheet("VES 2")
    bad.append(["Sounding Number", "VES 2"])
    bad.append(["No.", "AB/2 (m)", "MN (m)", "Apparent"])   # header, no rows
    empty = wb.create_sheet("Notes")
    empty.append(["Nothing here"])
    wb.save(path)
    del Workbook
    skipped = []
    soundings = read_ves_workbook(path, skipped=skipped)
    assert [s.sounding_id for s in soundings] == ["VES 1"]
    reasons = {f.message for f in skipped}
    assert any("'VES 2'" in r and "no numeric rows" in r for r in reasons)
    assert any("'Notes'" in r and "no data table" in r for r in reasons)
    assert all(f.code == "sheet_skipped" for f in skipped)
