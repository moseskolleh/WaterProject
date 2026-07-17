"""Costing module: rate catalogue, estimate roll-ups and the RWSN
worked-example numbers from "Costing and Pricing" (RWSN 2014-12)."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from groundwater.costing import (
    CostingInputs,
    RigSpec,
    annulus_volume_m3,
    estimate_borehole_cost,
    inputs_from_design,
    load_rates,
    loan_schedule,
    plot_cost_breakdown,
    rig_cost_per_well,
    running_cost_overburden_per_m,
    running_cost_rock_per_m,
    write_boq_workbook,
)
from groundwater.models import SiteMetadata
from groundwater.reporting.costing import CostReportInputs, build_cost_report


def test_rate_catalogue_loads_clean():
    rates = load_rates()
    assert len(rates) >= 20
    codes = [r.code for r in rates]
    assert len(codes) == len(set(codes)), "duplicate rate codes"
    for rate in rates:
        assert rate.unit_cost_usd > 0
        assert rate.category in ("equipment", "labour", "consumables", "fuel", "vehicles")


def test_estimate_rolls_up_consistently():
    est = estimate_borehole_cost(
        CostingInputs(total_depth_m=50, overburden_m=20, mobilisation_distance_km=100)
    )
    assert est.items, "no line items produced"
    by_stage = sum(v for _, v in est.by_stage())
    by_category = sum(v for _, v in est.by_category())
    assert by_stage == pytest.approx(est.direct_cost_usd)
    assert by_category == pytest.approx(est.direct_cost_usd)
    assert est.total_cost_usd == pytest.approx(est.direct_cost_usd * 1.15)
    assert est.price_usd == pytest.approx(est.total_cost_usd * 1.20)
    assert est.budget_usd == pytest.approx(est.price_usd * 1.10)
    assert est.cost_per_meter_usd == pytest.approx(est.total_cost_usd / 50)


def test_estimate_in_rwsn_worked_example_range():
    """The guide's 50 m example (20 m overburden, 100 km) costs
    6,466 USD (129.3 USD/m); the bundled indicative rates should land
    in the same range, not another order of magnitude."""
    est = estimate_borehole_cost(
        CostingInputs(total_depth_m=50, overburden_m=20, mobilisation_distance_km=100)
    )
    assert 80 <= est.cost_per_meter_usd <= 200
    assert 4_000 <= est.total_cost_usd <= 10_000


def test_vat_and_risk_loading():
    est = estimate_borehole_cost(CostingInputs(total_depth_m=50), vat_percent=15)
    assert est.price_with_vat_usd == pytest.approx(est.price_usd * 1.15)
    labels = [row[0] for row in est.summary_rows()]
    assert any("VAT" in label for label in labels)
    # no water no pay at 75 percent success carries a third more
    loaded = est.price_per_successful_well_usd(75)
    assert loaded == pytest.approx(est.price_with_vat_usd / 0.75)
    with pytest.raises(ValueError):
        est.price_per_successful_well_usd(0)


def test_zero_quantity_items_dropped_and_flagged_inputs():
    est = estimate_borehole_cost(CostingInputs(total_depth_m=50, handpumps=0))
    assert all(i.code != "PMP1" for i in est.items)
    assert any(f.code == "no_mobilisation_distance" for f in est.flags)
    assert est.assumptions, "rules of thumb should be recorded"


def test_estimate_rejects_nonpositive_depth():
    with pytest.raises(ValueError):
        estimate_borehole_cost(CostingInputs(total_depth_m=0))


def test_annulus_volume_matches_guide():
    """150 mm hole with 110 mm casing holds about 8 litres per metre."""
    v = annulus_volume_m3(150 / 25.4, 110 / 25.4, 1.0)
    assert v * 1000 == pytest.approx(8.2, abs=0.3)


def test_inputs_from_design_consistent(sample_data):
    from groundwater.design import design_borehole
    from groundwater.ingestion import read_drilling_workbook

    log = read_drilling_workbook(sample_data / "dr_timbo" / "dr_timbo_drilling_log.xlsx")
    design = design_borehole(log=log, static_water_level_m=10.0)
    inputs = inputs_from_design(design, mobilisation_distance_km=50)
    assert inputs.total_depth_m == design.total_depth_m
    assert inputs.screen_m == pytest.approx(design.total_screen_length_m)
    # casing + screen covers the hole plus stick-up
    assert inputs.casing_m + inputs.screen_m == pytest.approx(
        design.total_depth_m + design.stickup_m
    )


# ---------------------------------------------------------------------------
# Enterprise calculators: the guide's Table 3/4 and Box 2/4 numbers
# ---------------------------------------------------------------------------

def test_rig_depreciation_example():
    rig = RigSpec(capital_cost_usd=170_000)
    assert rig.depreciation_per_hour == pytest.approx(17.0)
    assert rig.depreciation_per_day == pytest.approx(136.0)
    assert rig.maintenance_per_hour == pytest.approx(0.85)


def test_wear_running_costs_example():
    assert running_cost_overburden_per_m() == pytest.approx(2.42, abs=0.01)
    assert running_cost_rock_per_m() == pytest.approx(7.42, abs=0.01)


def test_loan_schedule_example():
    """Box 2: 170,000 USD at 20 percent over 5 years is about 4,430
    USD per month (the guide used Excel PMT; allow 2 percent)."""
    loan = loan_schedule(170_000, 20, 5)
    assert loan.monthly_payment_usd == pytest.approx(4_430, rel=0.02)
    assert loan.total_paid_usd == pytest.approx(loan.monthly_payment_usd * 60)
    zero = loan_schedule(120_000, 0, 10)
    assert zero.monthly_payment_usd == pytest.approx(1_000)
    assert zero.total_interest_usd == pytest.approx(0)


def test_rig_cost_per_well_sensitivity():
    assert rig_cost_per_well(170_000, 5, 50) == pytest.approx(680)
    assert rig_cost_per_well(170_000, 5, 10) == pytest.approx(3_400)
    with pytest.raises(ValueError):
        rig_cost_per_well(170_000, 0, 10)


# ---------------------------------------------------------------------------
# Exports
# ---------------------------------------------------------------------------

def test_boq_workbook_and_report(tmp_path: Path):
    est = estimate_borehole_cost(
        CostingInputs(total_depth_m=65, mobilisation_distance_km=120)
    )
    boq = write_boq_workbook(est, tmp_path / "boq.xlsx")
    ws = load_workbook(boq).active
    assert ws["A1"].value.startswith("BILL OF QUANTITIES")
    # amounts are formulas so the workbook stays editable
    formulas = [c.value for row in ws.iter_rows(min_col=7, max_col=7) for c in row]
    assert any(isinstance(v, str) and v.startswith("=E") for v in formulas)

    plot_cost_breakdown(est, tmp_path / "breakdown.png")
    assert (tmp_path / "breakdown.png").stat().st_size > 10_000

    report = build_cost_report(
        CostReportInputs(
            estimate=est,
            site=SiteMetadata(community="Kuntolo", district="Bombali"),
            figures_dir=tmp_path,
        ),
        tmp_path / "cost.docx",
    )
    assert report.stat().st_size > 20_000


def test_report_is_reproducible(tmp_path: Path):
    est = estimate_borehole_cost(CostingInputs(total_depth_m=50))
    kwargs = dict(estimate=est, site=SiteMetadata(community="X"), figures_dir=tmp_path)
    a = build_cost_report(CostReportInputs(**kwargs), tmp_path / "a.docx")
    b = build_cost_report(CostReportInputs(**kwargs), tmp_path / "b.docx")
    assert a.read_bytes() == b.read_bytes()


def test_price_sensitivity_orders_and_brackets():
    from groundwater.costing import CostingInputs, price_sensitivity

    base, entries = price_sensitivity(
        CostingInputs(total_depth_m=60, mobilisation_distance_km=100)
    )
    assert base > 0 and entries
    spans = [e.span_usd for e in entries]
    assert spans == sorted(spans, reverse=True)
    depth_entry = next(e for e in entries if e.label.startswith("Total depth"))
    assert depth_entry.low_price_usd < base < depth_entry.high_price_usd
    labels = {e.label for e in entries}
    assert "Mobilisation distance ±50%" in labels
