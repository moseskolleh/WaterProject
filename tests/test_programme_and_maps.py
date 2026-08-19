"""Programme costing, regional maps, daily log template and the
metres reconciliation check."""

from __future__ import annotations

import numpy as np
import pytest
from openpyxl import load_workbook

from groundwater.costing import (
    CostingInputs,
    estimate_borehole_cost,
    estimate_programme_cost,
    plot_programme_gantt,
)
from groundwater.ingestion.templates import (
    write_all_templates,
    write_daily_log_template,
)
from groundwater.mapping import (
    load_admin,
    load_geology,
    load_hydrogeology,
    plot_admin_map,
    plot_geological_map,
    plot_hydrogeology_map,
)
from groundwater.models import SiteMetadata
from groundwater.reporting.context import context_map_figures
from groundwater.supervision import metres_reconciliation_check


# ---------------------------------------------------------------------------
# Programme costing
# ---------------------------------------------------------------------------

def _per_well() -> CostingInputs:
    return CostingInputs(total_depth_m=60, mobilisation_distance_km=150)


def test_programme_attempts_and_rollup():
    programme = estimate_programme_cost(
        _per_well(), 10, inter_site_distance_km=20, success_rate_percent=80
    )
    assert programme.n_attempted == 13  # ceil(10 / 0.8)
    well = programme.well_estimate
    # dry attempts cost less than complete wells
    assert 0 < programme.dry_attempt_cost_usd < well.direct_cost_usd
    expected_direct = (
        10 * well.direct_cost_usd
        + 3 * programme.dry_attempt_cost_usd
        + programme.transport_cost_usd
    )
    assert programme.direct_cost_usd == pytest.approx(expected_direct)
    assert programme.price_per_successful_well_usd == pytest.approx(
        programme.price_with_vat_usd / 10
    )


@pytest.mark.parametrize(
    "required,success,attempts",
    [
        (21, 35.0, 60),   # 60.000000000000007 in binary floating point
        (7, 5.6, 125),
        (14, 2.8, 500),
        (20, 35.0, 58),   # a genuine fraction still rounds up
        (10, 70.0, 15),
        (10, 100.0, 10),
    ],
)
def test_exact_attempt_counts_do_not_gain_a_phantom_dry_hole(
    required, success, attempts
):
    """21 wells at a 35 percent siting success rate is exactly 60 attempts.

    The division lands just above 60 in binary floating point, and ceil()
    then budgeted a 61st attempt - a whole dry borehole's drilling cost
    added to the programme.
    """
    programme = estimate_programme_cost(
        _per_well(), required, success_rate_percent=success
    )
    assert programme.n_attempted == attempts


def test_programme_transport_charged_once():
    """The package shares one mobilisation instead of one per well."""
    single = estimate_borehole_cost(_per_well())
    programme = estimate_programme_cost(
        _per_well(), 5, inter_site_distance_km=10, success_rate_percent=100
    )
    per_well_price = programme.price_with_vat_usd / 5
    assert per_well_price < single.price_usd
    # transport: 2 x 150 km base plus 4 moves of 10 km at the km rate
    km_rate = programme.transport_cost_usd / (2 * 150 + 4 * 10)
    assert km_rate > 0
    # the per-well estimate inside the programme carries no base transport
    assert programme.well_estimate.inputs.mobilisation_distance_km == 0


def test_programme_input_validation():
    with pytest.raises(ValueError):
        estimate_programme_cost(_per_well(), 0)
    with pytest.raises(ValueError):
        estimate_programme_cost(_per_well(), 5, success_rate_percent=0)


def test_programme_gantt(tmp_path):
    programme = estimate_programme_cost(_per_well(), 5)
    out = plot_programme_gantt(programme, tmp_path / "gantt.png")
    assert out.stat().st_size > 10_000


# ---------------------------------------------------------------------------
# Regional maps
# ---------------------------------------------------------------------------

def test_geology_layer_is_real_usgs_data():
    units = load_geology()
    assert len(units) >= 50
    codes = {u.glg for u in units}
    # the USGS units present in the Sierra Leone window
    assert {"pCm", "Qe", "Mi"} <= codes
    for unit in units:
        assert unit.color.startswith("#")
        ring = unit.ring
        assert ring.shape[1] == 2
        assert np.allclose(ring[0], ring[-1])  # closed
        assert (-13.6 < ring[:, 0]).all() and (ring[:, 0] < -10.0).all()
        assert (6.6 < ring[:, 1]).all() and (ring[:, 1] < 10.2).all()


def test_hydrogeology_layer_is_bgs_data():
    units = load_hydrogeology()
    codes = {u.glg for u in units}
    assert {"B-L", "U-M/H", "CSF-L/M", "I-L"} <= codes
    labels = {u.unit for u in units}
    assert any("productivity" in label for label in labels)


def test_admin_layer_is_geoboundaries_data():
    outline, districts = load_admin()
    assert outline.level == "ADM0" and outline.rings
    names = {d.name for d in districts}
    assert {"Bombali", "Bo", "Kambia", "Kenema"} <= names
    assert len(districts) == 14  # geoBoundaries release predates Karene/Falaba
    # district label points fall inside the country's bounding box
    for district in districts:
        lon, lat = district.label_point
        assert -13.6 < lon < -10.0 and 6.6 < lat < 10.1


def test_district_detection_from_coordinates():
    from groundwater.mapping import district_of

    assert district_of(9.03, -11.93) == "Bombali"
    assert district_of(8.47, -13.23) == "Western Area Urban"
    assert district_of(7.9, -11.2) == "Kenema"
    assert district_of(7.2, -13.4) == ""  # offshore


def test_maps_render(tmp_path):
    site = SiteMetadata(
        community="Kuntolo", district="Bombali",
        easting=178000, northing=1000000, utm_zone=29,
    )
    for name, fig_path in (
        ("national", plot_geological_map(site, path=tmp_path / "geo.png")),
        ("local", plot_geological_map(site, path=tmp_path / "geo_local.png",
                                      radius_km=40)),
        ("hydro", plot_hydrogeology_map(site, path=tmp_path / "hydro.png")),
        ("admin", plot_admin_map(site, path=tmp_path / "admin.png")),
    ):
        assert fig_path.stat().st_size > 30_000, name


def test_maps_render_without_site(tmp_path):
    plot_geological_map(None, path=tmp_path / "geo.png")
    plot_hydrogeology_map(None, path=tmp_path / "hydro.png")
    plot_admin_map(None, path=tmp_path / "admin.png")
    for name in ("geo.png", "hydro.png", "admin.png"):
        assert (tmp_path / name).exists()


def test_context_maps_for_reports(tmp_path):
    nowhere = SiteMetadata(community="X")
    assert context_map_figures(nowhere, tmp_path) == {}
    site = SiteMetadata(community="Kuntolo", district="Bombali",
                        easting=178000, northing=1000000, utm_zone=29)
    maps = context_map_figures(site, tmp_path)
    assert set(maps) == {"admin", "geology", "hydrogeology"}
    assert all(p.exists() for p in maps.values())


def test_context_maps_fall_back_to_the_recorded_area(tmp_path):
    """A site with no GPS fix still gets a map: of its district."""
    from groundwater.mapping import area_window

    site = SiteMetadata(community="Kuntoloh", district="Port Loko")
    window = area_window(site)
    assert window is not None
    assert window.exact is False
    assert window.label == "Port Loko district"

    maps = context_map_figures(site, tmp_path)
    assert set(maps) == {"admin", "geology", "hydrogeology"}
    assert all(p.exists() for p in maps.values())


def test_area_maps_reach_every_report_kind(tmp_path):
    """Every builder puts a map of the area in front of its numbers."""
    from docx import Document

    from groundwater.reporting.context import add_area_section
    from groundwater.reporting.docx_utils import ReportBuilder

    site = SiteMetadata(community="Kuntoloh", district="Port Loko")
    rb = ReportBuilder(None, title="T")
    maps = add_area_section(rb, site, tmp_path)
    assert set(maps) == {"admin", "geology", "hydrogeology"}
    out = tmp_path / "area.docx"
    rb.save(out)
    doc = Document(out)
    assert len(doc.inline_shapes) == 1
    text = "\n".join(p.text for p in doc.paragraphs)
    assert "Port Loko district" in text
    assert "No GPS position" in text


def test_handover_report_embeds_location_map(tmp_path):
    from groundwater.reporting.handover import (
        HandoverReportInputs,
        build_handover_report,
    )

    site = SiteMetadata(community="Kuntolo", district="Bombali",
                        easting=178000, northing=1000000, utm_zone=29)
    report = build_handover_report(
        HandoverReportInputs(site=site, figures_dir=tmp_path),
        tmp_path / "handover.docx",
    )
    assert list(tmp_path.glob("admin_map_*.png")), "location map not generated"
    assert report.stat().st_size > 100_000  # maps embedded


def test_context_maps_not_reused_across_sites(tmp_path):
    """Changing the coordinates must produce fresh maps, not reuse the
    previous site's figures."""
    site_a = SiteMetadata(community="A", easting=178000, northing=1000000,
                          utm_zone=29)
    site_b = SiteMetadata(community="B", easting=230000, northing=900000,
                          utm_zone=28)
    maps_a = context_map_figures(site_a, tmp_path)
    maps_b = context_map_figures(site_b, tmp_path)
    assert maps_a["admin"] != maps_b["admin"]
    assert maps_a["admin"].exists() and maps_b["admin"].exists()


# ---------------------------------------------------------------------------
# Daily log template and reconciliation
# ---------------------------------------------------------------------------

def test_daily_log_template(tmp_path):
    path = write_daily_log_template(tmp_path / "daily.xlsx")
    ws = load_workbook(path).active
    assert ws["A1"].value == "DRILLER'S DAILY REPORT"
    text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    for needed in ("Record taker", "Metres drilled today",
                   "Rig operator signature", "Supervisor signature"):
        assert needed in text


def test_all_templates_include_daily_log(tmp_path):
    paths = write_all_templates(tmp_path)
    names = {p.name for p in paths}
    assert "template_daily_drilling_report.xlsx" in names
    assert len(paths) == 5


def test_metres_reconciliation_boundaries():
    assert metres_reconciliation_check(60, 63).passed is True  # within 3 m
    over = metres_reconciliation_check(60, 64)
    assert over.passed is False and "withhold" in over.message
    under = metres_reconciliation_check(60, 50)
    assert under.passed is True and "covers all completed work" in under.message


def test_a_report_never_writes_figures_outside_the_directory_it_was_given(tmp_path,
                                                                         monkeypatch):
    """No builder may drop a PNG into the process working directory.

    The figures directory used to default to ``Path(".")``, so a report built
    without one scattered maps and charts wherever the process happened to be
    running - which is how three location maps ended up committed at the root
    of this repository.
    """
    from groundwater.costing import CostingInputs, estimate_borehole_cost
    from groundwater.reporting.costing import CostReportInputs, build_cost_report
    from groundwater.reporting.supervision import (
        SupervisionReportInputs,
        build_supervision_report,
    )
    from groundwater.supervision.checklists import (
        evaluate_checklist,
        load_checklists,
    )

    cwd = tmp_path / "cwd"
    cwd.mkdir()
    monkeypatch.chdir(cwd)
    out = tmp_path / "out"
    out.mkdir()

    site = SiteMetadata(community="Kuntolo", district="Bombali")
    items = load_checklists()
    build_supervision_report(
        SupervisionReportInputs(site=site, items=items, responses={},
                                assessment=evaluate_checklist(items, {})),
        out / "supervision.docx",
    )
    build_cost_report(
        CostReportInputs(
            estimate=estimate_borehole_cost(CostingInputs(total_depth_m=60)),
            site=site,
        ),
        out / "cost.docx",
    )

    assert list(cwd.iterdir()) == [], (
        "a report wrote into the working directory: "
        + ", ".join(p.name for p in cwd.iterdir())
    )
    assert (out / "supervision.docx").exists()
    assert any(p.suffix == ".png" for p in out.iterdir())
