"""Regression tests for the code-review bug fixes.

Each test pins the corrected behaviour of a specific defect found in the
project review so it cannot silently regress. Tests are grouped by the
subsystem they exercise.
"""

from __future__ import annotations

import numpy as np
import pytest
from openpyxl import Workbook

from groundwater.models import (
    LayeredModel,
    SiteMetadata,
    WaterQualityResult,
    WaterQualitySample,
    VESSounding,
)
from groundwater.quality import (
    assess_corrosivity,
    assess_health_risk,
    assess_sample,
    compute_wqi,
)


def _quality_sample(*results: WaterQualityResult) -> WaterQualitySample:
    return WaterQualitySample(site=SiteMetadata(community="Test"), results=list(results))


# --- water quality: microbiological classification --------------------------

def test_total_coliform_is_a_health_exceedance_not_aesthetic():
    """Faecal-indicator bacteria must never be reported as a taste problem."""
    a = assess_sample(_quality_sample(
        WaterQualityResult("pH", 7.2),
        WaterQualityResult("Total coliforms", 40.0, unit="CFU/100mL"),
        WaterQualityResult("Chloride", 30.0),
        WaterQualityResult("TDS", 200.0),
    ))
    names = {r.parameter for r in a.health_exceedances}
    assert "Total coliforms" in names
    assert "Total coliforms" not in {r.parameter for r in a.aesthetic_exceedances}
    assert "usable for drinking" not in a.verdict
    assert "health based guideline" in a.verdict


def test_ecoli_still_health_exceedance():
    a = assess_sample(_quality_sample(WaterQualityResult("E. coli", 12.0, unit="CFU/100mL")))
    assert "E. coli" in {r.parameter for r in a.health_exceedances}


# --- water quality: below-detection handling --------------------------------

def _quality_workbook(tmp_path, rows):
    """Write a minimal quality sheet: a header line + a results table."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Community:", "Test"])
    ws.append(["Sample ID:", "S1"])
    ws.append([])
    ws.append(["Parameter", "Unit", "Value", "Detection limit"])
    for row in rows:
        ws.append(row)
    path = tmp_path / "wq.xlsx"
    wb.save(path)
    return path


def test_below_detection_value_cleared_even_with_dl_column(tmp_path):
    """A '<X' value plus a filled detection-limit column must read as below
    detection, not as a real measurement equal to the limit."""
    from groundwater.ingestion import read_quality_workbook

    path = _quality_workbook(tmp_path, [
        ["Arsenic", "mg/L", "<0.01", "0.01"],   # marker AND explicit DL
        ["Lead", "mg/L", "<0.005", ""],           # marker only
    ])
    sample = read_quality_workbook(path)
    by_name = {r.parameter: r for r in sample.results}
    ars = by_name["Arsenic"]
    assert ars.below_detection is True
    assert ars.value is None
    assert ars.detection_limit == 0.01
    # and the assessment reports it as below-detection, not an exceedance
    a = assess_sample(sample)
    ars_row = next(r for r in a.rows if r.parameter == "Arsenic")
    assert ars_row.status == "below_detection"


# --- water quality: index sanity --------------------------------------------

def test_wqi_not_dominated_by_a_single_trace_toxicant():
    """A lone arsenic reading must not swamp the physico-chemical WQI."""
    params = dict(pH=7.0, Calcium=8.0, Magnesium=3.0, Sodium=12.0,
                  Chloride=15.0, Sulfate=6.0, TDS=90.0)
    clean = compute_wqi(_quality_sample(
        *[WaterQualityResult(k, v) for k, v in params.items()]))
    with_as = compute_wqi(_quality_sample(
        *[WaterQualityResult(k, v) for k, v in params.items()],
        WaterQualityResult("Arsenic", 0.02)))
    assert clean is not None and with_as is not None
    # adding one trace toxicant leaves the physico-chemical index unchanged
    assert with_as.value == clean.value
    assert clean.rating in ("Excellent", "Good")
    assert "Arsenic" not in {name for name, _ in with_as.top_contributors}


def test_nitrate_hazard_quotient_uses_nitrogen_basis():
    """Nitrate HQ must convert as-NO3 concentration to an as-N basis before
    dividing by the as-N reference dose (else it is ~4.4x too high)."""
    hr = assess_health_risk(_quality_sample(WaterQualityResult("Nitrate (as NO3)", 44.3)))
    assert hr is not None
    hq = hr.hazard_quotients["nitrate (as no3)"]
    # 44.3 as NO3 -> ~10 mg/L N; HQ = (10 * 2/70) / 1.6 ~= 0.18, well under the
    # ~0.79 the naive as-NO3 calculation would give.
    assert 0.15 < hq < 0.22


# --- water quality: corrosivity classification/flag consistency -------------

def test_corrosivity_label_matches_aggressive_flag():
    """When LSI/AI corroboration forces aggressive, the class label must not
    still read 'Balanced'/'Scale-forming'."""
    # soft, mildly acidic water: RSI lands near-balanced but LSI < -0.5
    corr = assess_corrosivity(_quality_sample(
        WaterQualityResult("pH", 6.6),
        WaterQualityResult("Calcium", 12.0),
        WaterQualityResult("Alkalinity", 40.0),
        WaterQualityResult("TDS", 120.0),
        WaterQualityResult("Temperature", 27.0),
    ))
    if corr.is_aggressive:
        assert corr.classification in ("Corrosive", "Strongly corrosive")
        assert "aggressive" in corr.verdict


# --- core: truthiness vs None -----------------------------------------------

def test_merged_with_keeps_real_zero_elevation():
    a = SiteMetadata(community="A", elevation_m=0.0)
    b = SiteMetadata(community="B", elevation_m=123.0)
    assert a.merged_with(b).elevation_m == 0.0  # a real 0 m is not "blank"
    # but a genuinely missing field is still filled
    c = SiteMetadata(community="C")
    assert c.merged_with(b).elevation_m == 123.0


def test_ves_segments_tolerate_blank_mn():
    """A blank (NaN) MN cell must not start a spurious one-point segment."""
    site = SiteMetadata(community="T")
    ab2 = np.array([1.5, 2.0, 3.0, 4.0, 6.0, 9.0])
    mn = np.array([0.5, 0.5, 0.5, np.nan, 5.0, 5.0])
    rho = np.array([100.0, 110.0, 120.0, 125.0, 130.0, 140.0])
    s = VESSounding(site=site, sounding_id="V1", ab2=ab2, mn=mn, rho_app=rho)
    segs = s.segments()
    assert len(segs) == 2  # the NaN folds into the running segment, not its own
    assert list(segs[0]) == [0, 1, 2, 3] and list(segs[1]) == [4, 5]
    # an entirely absent MN column collapses to one segment, not N of them
    s2 = VESSounding(site=site, sounding_id="V2", ab2=ab2, mn=np.full(6, np.nan),
                     rho_app=rho)
    assert len(s2.segments()) == 1


# --- portfolio: success rate cannot exceed 100% -----------------------------

def test_portfolio_success_rate_bounded():
    from groundwater.portfolio import portfolio_stats

    summaries = [
        {"status": "dry", "total_depth_m": 50},
        {"status": "successful, productive borehole"},   # no depth -> not drilled
        {"status": "successful"},                          # no depth -> not drilled
    ]
    stats = portfolio_stats(summaries)
    assert stats["n_drilled"] == 1
    assert stats["success_rate"] == 0.0  # the only drilled hole was dry
    assert stats["success_rate"] <= 100.0


# --- VES ingestion: blank spacer row must not truncate the table ------------

def test_ves_blank_spacer_row_does_not_truncate(tmp_path):
    from groundwater.ingestion.ves import read_ves_csv

    csv_text = (
        "No.,AB/2 (m),MN (m),rho (ohm-m)\n"
        "1,1.5,0.5,100\n"
        "2,2.0,0.5,110\n"
        "3,3.0,0.5,120\n"
        ",,,\n"              # blank spacer at the Schlumberger segment change
        "4,4.0,5,130\n"
        "5,6.0,5,140\n"
    )
    path = tmp_path / "ves.csv"
    path.write_text(csv_text)
    s = read_ves_csv(path)
    assert s is not None
    assert list(s.ab2) == [1.5, 2.0, 3.0, 4.0, 6.0]  # deep branch preserved
    # two consecutive blank rows still terminate the table
    csv_text2 = csv_text.replace(",,,\n4,4.0", ",,,\n,,,\n4,4.0")
    path2 = tmp_path / "ves2.csv"
    path2.write_text(csv_text2)
    s2 = read_ves_csv(path2)
    assert list(s2.ab2) == [1.5, 2.0, 3.0]


# --- VES: IPI2WIN fit-error parsing -----------------------------------------

def test_ipi2win_find_err_ignores_unrelated_words():
    from groundwater.ves.ipi2win import _find_err

    assert _find_err([["Terrain slope: 8 deg"]]) is None
    assert _find_err([["Field supervisor: Errol 5"]]) is None
    assert _find_err([["ERR = 3.5%"]]) == 3.5
    assert _find_err([["ERR (%): 21.5"]]) == 21.5


# --- VES: recommended drilling depth never exceeds the investigated depth ----

def test_max_drilling_depth_within_investigation():
    from groundwater.ves import interpret_model

    model = LayeredModel(np.array([300.0, 40.0]), np.array([6.0]))
    interp = interpret_model(None, model)
    assert interp.max_drilling_depth_m <= interp.investigation_depth_m


# --- design: geometry stays valid, dry intervals are not screened -----------

def test_dry_interval_not_screened():
    from groundwater.design import design_borehole
    from groundwater.models import DrillingLog, LithologyInterval

    log = DrillingLog(
        site=SiteMetadata(community="T"),
        total_depth_m=40.0,
        water_strikes_m=[],
        intervals=[
            LithologyInterval(0.0, 6.0, "laterite"),
            LithologyInterval(6.0, 30.0, "hard granite, dry, no water struck"),
            LithologyInterval(30.0, 40.0, "fractured granite, water-bearing"),
        ],
    )
    design = design_borehole(log=log, static_water_level_m=8.0)
    # the explicitly-dry interval must not be a screen target; the fractured
    # water-bearing interval must be
    assert any(s.top_m >= 30.0 - 1e-9 for s in design.segments if s.kind == "screen")
    assert not any(
        6.0 <= s.top_m < 30.0 and s.kind == "screen" for s in design.segments
    )


def test_shallow_hole_deep_swl_produces_valid_geometry():
    """A SWL too deep for the hole must not yield a negative-length screen or
    casing past the hole bottom - the geometry must stay valid and be flagged."""
    from groundwater.design import design_borehole
    from groundwater.models import DrillingLog

    log = DrillingLog(site=SiteMetadata(community="T"), total_depth_m=18.0,
                      water_strikes_m=[])
    design = design_borehole(log=log, static_water_level_m=14.0)
    for seg in design.segments:
        assert seg.bottom_m > seg.top_m           # no negative-length segment
        assert 0.0 <= seg.top_m <= design.total_depth_m + 1e-9
        assert seg.bottom_m <= design.total_depth_m + 1e-9
    assert design.total_screen_length_m > 0.0
    assert any(f.code == "hole_too_shallow" for f in design.flags)


# --- app: a loaded project with a non-string UTM zone must not brick it ------

def test_app_survives_integer_meta_zone():
    """A project file storing meta_zone as an int (e.g. 29) must not raise on
    every sidebar rerun."""
    from pathlib import Path

    streamlit = pytest.importorskip("streamlit")
    from streamlit.testing.v1 import AppTest

    app_path = str(Path(__file__).resolve().parents[1] / "app" / "streamlit_app.py")
    at = AppTest.from_file(app_path, default_timeout=600)
    at.session_state["meta_zone"] = 29  # int, as an older/hand-edited file might store
    at.run()
    assert not at.exception


# --- app: a deep sounding must not brick the guided start -------------------

def test_guided_start_survives_depth_beyond_the_widget_range():
    """A sounding that resolves no water zone recommends its investigated
    depth (max AB/2), which on a deep survey exceeds the guided start's
    300 m field. Streamlit raises on an out-of-range prefill, so the costing
    step took the whole page down with a red traceback."""
    from pathlib import Path

    pytest.importorskip("streamlit")
    from streamlit.testing.v1 import AppTest

    from groundwater.config import Config
    from groundwater.ves import interpret_model, invert_sounding
    from groundwater.ves.forward import forward_schlumberger

    cfg = Config()
    ab2 = np.array([1, 2, 3, 5, 7, 10, 15, 20, 30, 40, 60, 80, 100, 150,
                    200, 300, 400.0])
    # thin cover on fresh basement: nothing water bearing is resolved
    rho = forward_schlumberger((np.array([800.0, 4000.0]), np.array([4.0])), ab2)
    sounding = VESSounding(
        site=SiteMetadata(community="Dry", district="Bo"), sounding_id="D1",
        ab2=ab2, mn=ab2 / 3, rho_app=rho,
    )
    result = invert_sounding(sounding, cfg.ves)
    interp = interpret_model(sounding, result.model, cfg.ves)
    assert interp.max_drilling_depth_m > 300.0  # the precondition for the crash

    app_path = str(Path(__file__).resolve().parents[1] / "app" / "streamlit_app.py")
    at = AppTest.from_file(app_path, default_timeout=600)
    at.session_state["ves_results"] = ([sounding], [result], [interp])
    at.session_state["nav"] = "Guided start"
    at.session_state["wiz_step"] = 2  # the costing step
    at.run()
    assert not at.exception
    assert at.number_input(key="wiz_cost_depth").value <= 300.0


# --- app: checklist answers must survive moving between stages --------------

def test_supervision_answers_survive_a_stage_change():
    """Only the picked stage's widgets are drawn, and Streamlit discards the
    state of widgets a run does not draw. A supervisor who answered
    Procurement, moved to Drilling and came back found every answer reset to
    Pending - and the project file saved only the stage on screen."""
    from pathlib import Path

    pytest.importorskip("streamlit")
    from streamlit.testing.v1 import AppTest

    from groundwater.supervision import load_checklists

    stages: list[str] = []
    for item in load_checklists():
        if item.checklist not in stages:
            stages.append(item.checklist)
    assert len(stages) > 1

    app_path = str(Path(__file__).resolve().parents[1] / "app" / "streamlit_app.py")
    at = AppTest.from_file(app_path, default_timeout=600)
    at.session_state["nav"] = "Supervision"
    at.run()

    radios = [r for r in at.radio if r.key and r.key.startswith("chkw_")]
    assert radios, "no checklist items rendered"
    first, second = radios[0].key, radios[1].key
    item_id = first[len("chkw_"):]
    at.radio(key=first).set_value("Yes")
    at.run()
    at.radio(key=second).set_value("No")
    at.run()

    at.session_state["sup_stage"] = stages[-1]  # work a later stage
    at.run()
    # the answer lives in plain state, not in the widget that is no longer drawn
    assert at.session_state[f"chk_{item_id}"] == "Yes"

    at.session_state["sup_stage"] = stages[0]  # come back
    at.run()
    assert at.radio(key=first).value == "Yes"
    assert at.radio(key=second).value == "No"
    assert not at.exception


# --- ingestion: a blank test-type cell must not turn a step test constant ---

def test_blank_test_type_cell_keeps_the_steps(tmp_path):
    """The template's own title says "(STEP / CONSTANT DISCHARGE)".

    The banner scan read that as an answer, so a step test whose test-type
    cell was left blank came back as a constant-rate test with its three
    steps concatenated into one series - no well efficiency, no well-loss
    coefficients, and a transmissivity fitted to a discontinuous curve.
    """
    import shutil
    from pathlib import Path

    from openpyxl import load_workbook

    from groundwater.ingestion import read_pumping_workbook
    from groundwater.ingestion.pumping import _sheet_test_type

    # pre-printed form text is not an answer
    assert _sheet_test_type(
        [["PUMPING TEST FIELD SHEET (STEP / CONSTANT DISCHARGE)"]]
    ) == ""
    assert _sheet_test_type([["Constant discharge 61-120 min"]]) == ""
    # a genuine banner still is
    assert _sheet_test_type([["CONSTANT DISCHARGE TEST"]]) == "constant"
    assert _sheet_test_type([["STEP TEST"]]) == "step"

    src = (Path(__file__).resolve().parents[1] / "examples" / "data" /
           "kuntolo" / "kuntolo_step_test.xlsx")
    dst = tmp_path / "blank_type.xlsx"
    shutil.copy(src, dst)
    book = load_workbook(dst)
    sheet = book.active
    for row in range(1, 15):
        for col in range(1, 10):
            value = sheet.cell(row, col).value
            if isinstance(value, str) and "test type" in value.lower():
                sheet.cell(row, col + 1).value = None
    book.save(dst)

    test = read_pumping_workbook(dst)
    assert test.test_type.startswith("step")
    assert len(test.steps) == 3
    # and the guess is surfaced rather than made silently
    assert any(f.code == "test_type_inferred" for f in test.flags)


# --- hydraulics: a zero discharge is a blank, not a rate --------------------

def test_zero_discharge_is_treated_as_not_measured():
    """A 0 typed into the discharge row divided by zero deep inside the
    Cooper-Jacob fit. Transmissivity is proportional to Q, so a zero rate is
    a blank the crew wrote a number into - it belongs on the same handled
    "pending" path as a missing value, not in a traceback."""
    from groundwater.config import Config
    from groundwater.hydraulics import analyse_pumping_test
    from groundwater.hydraulics.analysis import cooper_jacob
    from groundwater.models import PumpingStep, PumpingTest

    time_min = np.array([1.0, 2, 3, 5, 8, 12, 20, 30, 45, 60])
    level = 10 + 1.5 * np.log10(time_min)

    for bad in (0.0, -2.0, float("nan")):
        test = PumpingTest(
            site=SiteMetadata(community="Z"), test_type="constant",
            steps=[PumpingStep(step_number=1, time_min=time_min,
                               water_level_m=level, discharge_m3_per_h=bad)],
            static_water_level_m=10.0, borehole_depth_m=50.0,
        )
        analysis = analyse_pumping_test(test, Config().pumping)
        assert analysis.transmissivity_m2_per_day is None
        assert any(f.code == "invalid_discharge" for f in analysis.flags)

    # a real rate still analyses
    test = PumpingTest(
        site=SiteMetadata(community="Z"), test_type="constant",
        steps=[PumpingStep(step_number=1, time_min=time_min,
                           water_level_m=level, discharge_m3_per_h=2.93)],
        static_water_level_m=10.0, borehole_depth_m=50.0,
    )
    assert analyse_pumping_test(test, Config().pumping).transmissivity_m2_per_day > 0

    # and calling a fit directly gives a clear error, not ZeroDivisionError
    with pytest.raises(ValueError, match="greater than zero"):
        cooper_jacob(time_min, level - 10, 0.0)


def test_interpret_model_handles_a_bare_half_space():
    """A single-layer model has no finite layer bottom; the investigated
    depth fell back to bottoms[-2] and raised IndexError."""
    from groundwater.ves import interpret_model

    model = LayeredModel(resistivities=np.array([250.0]), thicknesses=np.array([]))
    interp = interpret_model(None, model)
    assert len(interp.layers) == 1
    # nothing was resolved, so nothing is recommended
    assert interp.max_drilling_depth_m == 0.0


def test_discharges_do_not_carry_over_to_another_borehole():
    """The discharge boxes are keyed by step number, so they outlived the
    sheet they were typed for. A second borehole whose sheet also lacks
    discharges silently inherited the first one's rates - and transmissivity,
    safe yield and pump setting depth are all proportional to them."""
    from pathlib import Path

    pytest.importorskip("streamlit")
    from streamlit.testing.v1 import AppTest

    app_path = str(Path(__file__).resolve().parents[1] / "app" / "streamlit_app.py")
    at = AppTest.from_file(app_path, default_timeout=600)
    at.session_state["nav"] = "Pumping test"
    # rates left behind by a previously open sheet
    for step in (1, 2, 3):
        at.session_state[f"q_{step}"] = 9.9
    at.run()
    at.selectbox(key="sample_pump").select("kuntolo/kuntolo_step_test.xlsx")
    at.run()
    assert not at.exception
    analysis = at.session_state["pump_analysis"]
    assert [s.discharge_m3_per_h for s in analysis.test.steps] == [None, None, None]

    # a rate typed for this sheet is still used
    for step, value in ((1, 2.5), (2, 3.5), (3, 4.5)):
        at.number_input(key=f"q_{step}").set_value(value)
        at.run()
    analysis = at.session_state["pump_analysis"]
    assert [s.discharge_m3_per_h for s in analysis.test.steps] == [2.5, 3.5, 4.5]


def test_excel_date_in_the_depth_column_is_rejected_and_reported(tmp_path):
    """Excel turns "5-10" typed into a General cell into 10 May.

    openpyxl hands that back as a datetime, and reading it as text gave a
    lithology interval of 5 to 2026 metres - which then placed screens and
    priced casing against a two-kilometre hole.
    """
    import datetime
    import shutil
    from pathlib import Path

    from openpyxl import load_workbook

    from groundwater.ingestion import read_drilling_workbook
    from groundwater.utils import parse_depth_interval

    assert parse_depth_interval(datetime.datetime(2026, 5, 10)) is None
    assert parse_depth_interval(datetime.date(2026, 1, 5)) is None
    assert parse_depth_interval("5-10") == (5.0, 10.0)  # still parses

    src = (Path(__file__).resolve().parents[1] / "examples" / "data" /
           "dr_timbo" / "dr_timbo_drilling_log.xlsx")
    dst = tmp_path / "date_interval.xlsx"
    shutil.copy(src, dst)
    book = load_workbook(dst)
    sheet = book.active
    converted = 0
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == "5-10":
                cell.value = datetime.datetime(2026, 5, 10)
                converted += 1
    assert converted, "sample log has no 5-10 interval to convert"
    book.save(dst)

    log = read_drilling_workbook(dst)
    assert max(i.bottom_m for i in log.intervals) < 200  # no 2026 m interval
    assert any(f.code == "interval_read_as_date" for f in log.flags)


def test_two_boreholes_do_not_share_report_figures(tmp_path):
    """Report figures were written under fixed filenames and regenerated only
    when absent. The app gives every report in a session the same figures
    directory, so the second borehole's client report embedded the first
    borehole's water-level record.
    """
    import hashlib
    from pathlib import Path

    from groundwater.config import Config
    from groundwater.hydraulics import analyse_pumping_test
    from groundwater.ingestion import read_pumping_workbook
    from groundwater.reporting.pumping import (
        PumpingReportInputs,
        build_pumping_report,
    )

    data = Path(__file__).resolve().parents[1] / "examples" / "data"
    config = Config()

    first = read_pumping_workbook(data / "dr_timbo" / "dr_timbo_constant_test.xlsx")
    build_pumping_report(
        PumpingReportInputs(analysis=analyse_pumping_test(first, config.pumping),
                            figures_dir=tmp_path),
        out_path=tmp_path / "first.docx", config=config,
    )
    second = read_pumping_workbook(data / "kuntolo" / "kuntolo_step_test.xlsx")
    for step in second.steps:
        step.discharge_m3_per_h = 2.5
    build_pumping_report(
        PumpingReportInputs(analysis=analyse_pumping_test(second, config.pumping),
                            figures_dir=tmp_path),
        out_path=tmp_path / "second.docx", config=config,
    )

    overviews = sorted(tmp_path.glob("test_overview*.png"))
    assert len(overviews) == 2, "both boreholes must get their own figure"
    digests = {hashlib.sha256(p.read_bytes()).hexdigest() for p in overviews}
    assert len(digests) == 2, "the second report reused the first one's figure"


def test_a_national_standard_failure_is_not_reported_as_a_taste_problem():
    """A national limit can be stricter than the WHO health guideline.

    Exceeding it was folded into the aesthetic bucket, and the report then
    said the water "is usable for drinking" and only warned about taste -
    for a supply that fails the national standard. QUESTIONS.md asks the
    user to replace the national column with the real Standards Bureau
    values, so this is the column most likely to tighten.
    """
    # Aluminium: WHO health 0.9 mg/L, national 0.2 mg/L
    strict = assess_sample(_quality_sample(
        WaterQualityResult("pH", 7.2),
        WaterQualityResult("Aluminium", 0.5),
    ))
    assert [r.parameter for r in strict.national_exceedances] == ["Aluminium"]
    assert not strict.health_exceedances
    assert "does not comply with the national standard" in strict.verdict
    assert "usable for drinking" not in strict.verdict

    # Iron has no WHO health value at all, so its limit is an acceptability
    # one and the friendly wording is right
    taste = assess_sample(_quality_sample(
        WaterQualityResult("pH", 7.2),
        WaterQualityResult("Iron", 0.5),
    ))
    assert not taste.national_exceedances
    assert "usable for drinking" in taste.verdict


def test_saving_right_after_an_analysis_captures_it():
    """The sidebar renders before every page body, so a "Save project"
    button built there carried the state as it was *before* this run's
    analyses. Saving straight after running the siting wrote a file with no
    results in it, and the Portfolio page then showed the site as unstarted.
    The panel is filled at the end of the script instead.

    The same ordering left interp.rank unset when the Overview read it - only
    the VES page ranked, and it renders later - so the dashboard named
    whichever sounding was parsed first as the preferred drill target.
    """
    from pathlib import Path

    pytest.importorskip("streamlit")
    from streamlit.testing.v1 import AppTest

    app_path = str(Path(__file__).resolve().parents[1] / "app" / "streamlit_app.py")
    at = AppTest.from_file(app_path, default_timeout=600)
    at.run()
    at.text_input(key="meta_community").set_value("Rokel")
    at.run()
    at.selectbox(key="sample_ves").select("rokel/rokel_ves.xlsx")
    at.run()
    at.button(key="run_ves").click()
    at.run()  # the run that produces the results
    assert not at.exception

    summary = at.session_state["project_summary"]
    assert summary["status"], "the saved project must carry the analysis just run"

    _, _, interps = at.session_state["ves_results"]
    assert all(i.rank for i in interps), "ranks must be set where interps are built"
    best = min(interps, key=lambda i: (i.rank or 99, -i.score))
    assert best.score == max(i.score for i in interps)


# --- robustness one-liners --------------------------------------------------

def test_loan_schedule_rejects_zero_term():
    from groundwater.costing.enterprise import loan_schedule

    with pytest.raises(ValueError):
        loan_schedule(10000, 5, 0)
    # a normal loan still works
    summ = loan_schedule(10000, 10, 5)
    assert summ.monthly_payment_usd > 0


def test_site_location_map_empty_points_clear_error(tmp_path):
    from groundwater.mapping import site_location_map

    with pytest.raises(ValueError):
        site_location_map([], zone=28, path=tmp_path / "x.png")


def test_coordinate_flag_uses_west_hemisphere():
    """Sierra Leone longitudes are negative and must read 'W', not 'E'."""
    from groundwater.ingestion.checks import _fmt_latlon

    text = _fmt_latlon(8.5859, -12.4562)
    assert text == "8.5859 N, 12.4562 W"
