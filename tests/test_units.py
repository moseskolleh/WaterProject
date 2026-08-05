"""Units are read, not assumed.

Every number this toolkit reads arrives beside a unit written by hand. Until
these tests existed, those units were decoration: a laboratory result was
compared against a guideline whatever unit each carried, and a pumping
discharge was taken as m3/h whatever the sheet said. Both are silent,
order-of-magnitude errors in a safety judgement, so each one has a test
naming the failure it prevents.
"""

from __future__ import annotations

import pytest

from groundwater.models import (
    SiteMetadata,
    WaterQualityResult,
    WaterQualitySample,
)
from groundwater.quality import (
    assess_corrosivity,
    assess_health_risk,
    assess_sample,
    compute_wqi,
    ionic_balance,
)
from groundwater.units import (
    canonical_unit,
    convert,
    parse_unit,
    read_quantity,
    unit_from_label,
)


def _sample(*results: WaterQualityResult) -> WaterQualitySample:
    return WaterQualitySample(site=SiteMetadata(community="Test"), results=list(results))


# --- the conversion table ----------------------------------------------------

@pytest.mark.parametrize("value,source,target,expected", [
    (5.0, "ug/L", "mg/L", 0.005),
    (5.0, "µg/L", "mg/L", 0.005),        # MICRO SIGN
    (5.0, "μg/L", "mg/L", 0.005),        # GREEK SMALL LETTER MU
    (5.0, "ppb", "mg/L", 0.005),
    (5.0, "ppm", "mg/L", 5.0),
    (0.5, "g/L", "mg/L", 500.0),
    (0.185, "mS/cm", "uS/cm", 185.0),
    (2.0, "CFU/mL", "CFU/100 mL", 200.0),
    (2.0, "L/s", "m3/h", 7.2),
    (120.0, "L/min", "m3/h", 7.2),
    (1.0, "m3/d", "m3/h", 1 / 24),
    (2.0, "hours", "min", 120.0),
    (60.0, "s", "min", 1.0),
])
def test_known_units_convert(value, source, target, expected):
    got = convert(value, source, target)
    assert got == pytest.approx(expected)


@pytest.mark.parametrize("source,target", [
    ("mg/L", "NTU"),          # different dimensions
    ("squiggles", "mg/L"),    # not a unit at all
    ("gpm", "m3/h"),          # US or imperial gallon? refused, not guessed
    ("ppt", "mg/L"),          # parts per thousand or per trillion?
    ("JTU", "NTU"),           # not numerically interchangeable with NTU
    ("mg/L as N", "mg/L as NO3"),   # same scale, different chemistry
])
def test_units_that_must_be_refused(source, target):
    """None means "do not use this number". It never means zero."""
    assert convert(1.0, source, target) is None


def test_a_bare_letter_is_read_by_the_column_it_is_in():
    assert parse_unit("m", dimension="time").canonical == "min"
    assert parse_unit("m", dimension="length").canonical == "m"
    assert parse_unit("mg/L", dimension="time") is None


def test_canonical_spelling():
    assert canonical_unit("µS/cm") == "uS/cm"
    assert canonical_unit("LPS") == "L/s"
    assert canonical_unit("nonsense") == "nonsense"


# --- reading a unit off a sheet ---------------------------------------------

def test_only_bracketed_text_counts_as_a_declared_unit():
    """So an ordinary word in a label is never reported as an unknown unit."""
    assert unit_from_label("Time (min)", dimension="time")[0] == "min"
    assert unit_from_label("Time (h)", dimension="time")[1].factor == 60.0
    assert unit_from_label("Discharge per step (L/s)", dimension="flow")[0] == "L/s"
    assert unit_from_label("Step 1 Q", dimension="flow") == ("", None)
    assert unit_from_label("Time", dimension="time") == ("", None)
    # prose in brackets is not a unit
    assert unit_from_label("Constant discharge (0-60 min)", dimension="time")[1] is None
    # a declared unit that cannot be read is reported as declared, not absent
    written, unit = unit_from_label("Time (furlongs)", dimension="time")
    assert written == "furlongs" and unit is None


def test_the_cell_outranks_its_header():
    """A crew that types "0.81 L/s" into a column headed m3/h means L/s."""
    q = read_quantity("0.81 L/s", "Discharge per step (m3/h)", dimension="flow")
    assert q.status == "converted"
    assert q.value == pytest.approx(2.916)


def test_an_unreadable_unit_yields_no_value_at_all():
    q = read_quantity(2.0, "Q (gpm)", dimension="flow")
    assert q.status == "unknown"
    assert q.value is None          # never a fallback to the raw number
    assert q.raw_value == 2.0


def test_no_declared_unit_assumes_the_canonical_one_and_says_so():
    q = read_quantity(2.5, "Step 1 Q", dimension="flow")
    assert q.status == "assumed" and q.value == 2.5


# --- what the units protect --------------------------------------------------

def test_a_microgram_result_is_not_graded_as_milligrams():
    """5 ug/L of arsenic is half the guideline; read as mg/L it is 500x over.

    The same arithmetic in the other direction certifies unsafe water as
    safe, which is why this is a safety fix and not a tidiness one.
    """
    micro = assess_sample(_sample(WaterQualityResult("Arsenic", 5.0, "ug/L")))
    assert micro.rows[0].status == "within_limits"
    assert micro.rows[0].value_in_guideline_unit == pytest.approx(0.005)

    milli = assess_sample(_sample(WaterQualityResult("Arsenic", 5.0, "mg/L")))
    assert milli.rows[0].status == "exceeds_health"


def test_a_gram_per_litre_nitrate_is_not_read_as_compliant():
    """The dangerous direction: 0.1 g/L is 100 mg/L, twice the guideline."""
    a = assess_sample(_sample(WaterQualityResult("Nitrate (as NO3)", 0.1, "g/L")))
    assert a.rows[0].status == "exceeds_health"


def test_conductivity_in_millisiemens_is_converted():
    a = assess_sample(_sample(
        WaterQualityResult("Electrical conductivity", 3.0, "mS/cm")))
    assert a.rows[0].value_in_guideline_unit == pytest.approx(3000.0)
    assert a.rows[0].status == "exceeds_aesthetic"


def test_an_unreadable_unit_makes_the_row_indeterminate_not_compliant():
    a = assess_sample(_sample(WaterQualityResult("Iron", 0.1, "wibbles")))
    assert a.rows[0].status == "indeterminate"
    assert a.rows[0].reason == "unit_unreadable"
    assert a.verdict_state == "indeterminate"


def test_a_unit_from_another_dimension_is_refused():
    a = assess_sample(_sample(WaterQualityResult("Iron", 0.1, "NTU")))
    assert a.rows[0].status == "indeterminate"
    assert a.rows[0].reason == "unit_mismatch"


def _chemistry(unit: str, scale: float) -> WaterQualitySample:
    rows = {
        "TDS": 120.0, "Chloride": 18.4, "Sulfate": 6.8,
        "Nitrate (as NO3)": 8.2, "Iron": 0.05, "Fluoride": 0.5,
        "Arsenic": 0.01, "Calcium": 14.2, "Magnesium": 5.5, "Sodium": 12.6,
        "Potassium": 2.1, "Bicarbonate": 56.0, "Alkalinity": 46.0,
    }
    results = [
        WaterQualityResult("pH", 7.2, "pH units"),
        WaterQualityResult("Temperature", 27.4, "deg C"),
    ]
    results += [WaterQualityResult(k, v * scale, unit) for k, v in rows.items()]
    return _sample(*results)


@pytest.mark.parametrize("unit,scale", [
    ("ug/L", 1000.0), ("ppm", 1.0), ("g/L", 0.001),
])
def test_the_aggregate_indices_do_not_depend_on_the_reported_unit(unit, scale):
    """The same water, written three ways, is the same water.

    The reference doses and index weights are all mg/L figures, so a sample
    reported in ug/L used to produce a hazard index a thousand times too
    high - "High" risk, a 0.4 lifetime cancer risk and an "Unsuitable for
    drinking" quality index - for water well inside every guideline.
    """
    base, other = _chemistry("mg/L", 1.0), _chemistry(unit, scale)
    assert assess_health_risk(other).hazard_index == pytest.approx(
        assess_health_risk(base).hazard_index)
    assert compute_wqi(other).value == pytest.approx(compute_wqi(base).value)
    assert ionic_balance(other).error_percent == pytest.approx(
        ionic_balance(base).error_percent)
    assert assess_corrosivity(other).rsi == pytest.approx(
        assess_corrosivity(base).rsi)


# --- pumping sheets ----------------------------------------------------------

def _pumping_copy(tmp_path, source, edits):
    """A copy of a sample pumping workbook with a few cells rewritten."""
    import shutil

    from openpyxl import load_workbook

    path = tmp_path / "sheet.xlsx"
    shutil.copy(source, path)
    wb = load_workbook(path)
    ws = wb.active
    for cell, value in edits.items():
        ws[cell] = value
    wb.save(path)
    return path


@pytest.fixture
def constant_test(sample_data):
    return sample_data / "dr_timbo" / "dr_timbo_constant_test.xlsx"


def test_a_discharge_in_litres_per_second_is_converted(tmp_path, constant_test):
    """0.814 L/s is 2.93 m3/h. Read as m3/h it understates the yield 3.6x,
    and every transmissivity and pump setting derived from it with it."""
    from groundwater.ingestion import read_pumping_workbook

    path = _pumping_copy(tmp_path, constant_test,
                         {"A9": "Discharge per step (L/s)", "C9": 0.814})
    test = read_pumping_workbook(path)
    assert test.steps[0].discharge_m3_per_h == pytest.approx(2.9304)
    assert any(f.code == "discharge_unit_converted" for f in test.flags)


def test_an_unreadable_discharge_unit_leaves_the_step_pending(tmp_path, constant_test):
    """Refusing routes into the existing "results pending" path, which is a
    far better answer than a number in the wrong unit."""
    from groundwater.ingestion import read_pumping_workbook

    path = _pumping_copy(tmp_path, constant_test,
                         {"A9": "Discharge per step (gpm)", "C9": 12.9})
    test = read_pumping_workbook(path)
    assert test.steps[0].discharge_m3_per_h is None
    assert any(f.code == "discharge_unit_unknown" for f in test.flags)
    assert any(f.code == "missing_discharge" for f in test.flags)


def test_times_recorded_in_hours_are_converted(tmp_path, constant_test):
    from groundwater.ingestion import read_pumping_workbook

    edits = {f"{col}11": "Time (h)" for col in ("A", "D", "G", "J", "M")}
    path = _pumping_copy(tmp_path, constant_test, edits)
    test = read_pumping_workbook(path)
    assert list(test.steps[0].time_min[:3]) == [0.0, 60.0, 120.0]
    assert any(f.code == "time_unit_converted" for f in test.flags)


def test_an_unreadable_time_unit_drops_the_column(tmp_path, constant_test):
    """There is no pending state for time: every consumer would silently fit
    a curve to the wrong axis, so the readings are refused instead."""
    from groundwater.hydraulics import analyse_pumping_test
    from groundwater.ingestion import read_pumping_workbook

    edits = {f"{col}11": "Time (furlongs)" for col in ("A", "D", "G", "J", "M")}
    path = _pumping_copy(tmp_path, constant_test, edits)
    test = read_pumping_workbook(path)
    assert test.steps == []
    assert test.recovery_time_min is None
    errors = [f for f in test.flags if f.code == "time_unit_unknown"]
    assert errors and all(f.level == "error" for f in errors)
    assert "furlongs" in errors[0].message
    # and the analysis carries the reason forward rather than inventing a
    # transmissivity from nothing
    analysis = analyse_pumping_test(test)
    assert analysis.transmissivity_m2_per_day is None
    assert any(f.code == "time_unit_unknown" for f in analysis.flags)


def test_the_next_steps_label_is_not_read_as_this_steps_discharge():
    """Scanning the neighbouring cell blindly read "Step 2 Q (m3/h)" as a
    discharge of 2 m3/h whenever the first step's box was left empty, and
    fitted a transmissivity to it."""
    from groundwater.ingestion.pumping import _find_step_discharges

    assert _find_step_discharges([["Step 1 Q (m3/h)", "", "Step 2 Q (m3/h)", ""]]) == {}
    filled = _find_step_discharges(
        [["Discharge per step (m3/h)", "Step 1 Q", 2.5, "", "Step 2 Q", 3.0]])
    assert filled[1].value == 2.5 and filled[2].value == 3.0


def test_the_sample_sheets_still_read_exactly_as_before(sample_data):
    """The units fix must not disturb a sheet that already used m3/h."""
    from groundwater.ingestion import read_pumping_workbook

    test = read_pumping_workbook(
        sample_data / "dr_timbo" / "dr_timbo_constant_test.xlsx")
    assert test.steps[0].discharge_m3_per_h == pytest.approx(2.93)
    assert list(test.steps[0].time_min[:4]) == [0.0, 1.0, 2.0, 3.0]
    assert not [f for f in test.flags if "unit" in f.code]
